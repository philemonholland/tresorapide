from datetime import date
from decimal import Decimal
from django.test import TestCase, RequestFactory
from django.contrib.sessions.middleware import SessionMiddleware
from django.core.files.uploadedfile import SimpleUploadedFile
from django.urls import reverse

from accounts.models import User
from houses.models import House
from budget.models import BudgetYear, SubBudget, Expense
from budget.services import BudgetCalculationService
from budget.views import BudgetYearCreateView
from bons.models import BonDeCommande, BonStatus, ReceiptFile, OcrStatus
from members.models import Member, Apartment, Residency


class BudgetYearAutoContingencyTests(TestCase):
    def test_imprevues_sub_budget_auto_created(self):
        house = House.objects.create(code="BB", name="Maison BB", account_number="13-51200")
        by = BudgetYear.objects.create(
            house=house, year=2026, annual_budget_total=Decimal("12237.00")
        )
        imprevues = by.sub_budgets.filter(trace_code=0).first()
        self.assertIsNotNone(imprevues)
        self.assertTrue(imprevues.is_contingency)
        self.assertEqual(imprevues.planned_amount, Decimal("12237.00") * Decimal("0.15"))
        self.assertEqual(imprevues.name, "Imprévues")


class BudgetCalculationTests(TestCase):
    def setUp(self):
        self.house = House.objects.create(code="BB", name="Maison BB", account_number="13-51200")
        self.by = BudgetYear.objects.create(
            house=self.house, year=2026,
            annual_budget_total=Decimal("12237.00"),
            snow_budget=Decimal("1860.00"),
        )
        # Get the auto-created imprévues sub-budget
        self.sb_imprevues = self.by.sub_budgets.get(trace_code=0)
        self.sb_repairs_apt = SubBudget.objects.create(
            budget_year=self.by, trace_code=1,
            name="Réparations par appartement", planned_amount=Decimal("2000.00")
        )
        self.sb_repairs_bb = SubBudget.objects.create(
            budget_year=self.by, trace_code=2,
            name="Réparations BB", planned_amount=Decimal("1000.00")
        )
        self.sb_produits = SubBudget.objects.create(
            budget_year=self.by, trace_code=7,
            name="Produits ménager", planned_amount=Decimal("300.00")
        )

    def test_base_values_no_expenses(self):
        vals = BudgetCalculationService.base_values(self.by)
        self.assertEqual(vals["budget_total"], Decimal("12237.00"))
        self.assertEqual(vals["snow_budget"], Decimal("1860.00"))
        self.assertEqual(vals["imprevues"], Decimal("12237.00") * Decimal("0.15"))
        self.assertEqual(vals["expenses_to_date"], Decimal("0"))

    def test_base_values_with_expenses(self):
        Expense.objects.create(
            budget_year=self.by, sub_budget=self.sb_produits,
            entry_date=date(2026, 1, 7), description="Nettoyants",
            amount=Decimal("19.49"), spent_by_label="202 / Marylin"
        )
        vals = BudgetCalculationService.base_values(self.by)
        self.assertEqual(vals["expenses_to_date"], Decimal("19.49"))

    def test_repair_totals(self):
        Expense.objects.create(
            budget_year=self.by, sub_budget=self.sb_repairs_apt,
            entry_date=date(2026, 2, 26), description="Robinet",
            amount=Decimal("54.54"), spent_by_label="304 / Serge"
        )
        totals = BudgetCalculationService.repair_totals(self.by)
        self.assertEqual(totals["planned"], Decimal("3000.00"))
        self.assertEqual(totals["used"], Decimal("54.54"))
        self.assertEqual(totals["remaining"], Decimal("2945.46"))

    def test_imprevues_totals(self):
        Expense.objects.create(
            budget_year=self.by, sub_budget=self.sb_imprevues,
            entry_date=date(2026, 1, 30), description="Pouliot repair",
            amount=Decimal("564.92"), spent_by_label="BB",
            source_type="accountant_direct"
        )
        totals = BudgetCalculationService.imprevues_totals(self.by)
        self.assertEqual(totals["used"], Decimal("564.92"))
        expected_remaining = Decimal("12237.00") * Decimal("0.15") - Decimal("564.92")
        self.assertEqual(totals["remaining"], expected_remaining)

    def test_available_money(self):
        Expense.objects.create(
            budget_year=self.by, sub_budget=self.sb_produits,
            entry_date=date(2026, 1, 7), description="Nettoyants",
            amount=Decimal("19.49"), spent_by_label="202 / Marylin"
        )
        Expense.objects.create(
            budget_year=self.by, sub_budget=self.sb_imprevues,
            entry_date=date(2026, 1, 30), description="Pouliot",
            amount=Decimal("564.92"), spent_by_label="BB",
            source_type="accountant_direct"
        )
        avail = BudgetCalculationService.available_money(self.by)
        self.assertEqual(avail["available"], Decimal("12237.00") - Decimal("19.49") - Decimal("564.92"))

    def test_category_summary(self):
        Expense.objects.create(
            budget_year=self.by, sub_budget=self.sb_produits,
            entry_date=date(2026, 1, 7), description="Nettoyants",
            amount=Decimal("19.49"), spent_by_label="202 / Marylin"
        )
        summary = BudgetCalculationService.category_summary(self.by)
        produits = [s for s in summary if s["trace_code"] == 7][0]
        self.assertEqual(produits["used"], Decimal("19.49"))
        self.assertEqual(produits["remaining"], Decimal("300.00") - Decimal("19.49"))

    def test_running_balances(self):
        Expense.objects.create(
            budget_year=self.by, sub_budget=self.sb_produits,
            entry_date=date(2026, 1, 7), description="Nettoyants",
            amount=Decimal("19.49"), spent_by_label="202 / Marylin"
        )
        Expense.objects.create(
            budget_year=self.by, sub_budget=self.sb_imprevues,
            entry_date=date(2026, 1, 30), description="Pouliot",
            amount=Decimal("564.92"), spent_by_label="BB"
        )
        balances = BudgetCalculationService.running_balances(self.by)
        self.assertEqual(len(balances), 2)
        self.assertEqual(balances[0]["balance"], Decimal("12237.00") - Decimal("19.49"))
        self.assertEqual(balances[1]["balance"], Decimal("12237.00") - Decimal("19.49") - Decimal("564.92"))
        budget_minus_15 = Decimal("12237.00") - Decimal("12237.00") * Decimal("0.15")
        self.assertEqual(balances[0]["balance_minus_imprevues"], budget_minus_15 - Decimal("19.49"))
        self.assertEqual(balances[1]["balance_minus_imprevues"], budget_minus_15 - Decimal("19.49"))

    def test_unbudgeted_available(self):
        vals = BudgetCalculationService.unbudgeted_available(self.by)
        # budget_total=12237, imprevues=12237*0.15=1835.55
        # planned non-contingency: 2000 + 1000 + 300 = 3300
        # unbudgeted = 12237 - 3300 = 8937 (no imprevues subtraction)
        expected = Decimal("12237.00") - Decimal("3300.00")
        self.assertEqual(vals["unbudgeted_available"], expected)
        # minus_15 = (12237 - imprevues) - 3300
        imprevues = Decimal("12237.00") * Decimal("0.15")
        expected_minus_15 = Decimal("12237.00") - imprevues - Decimal("3300.00")
        self.assertEqual(vals["unbudgeted_available_minus_15"], expected_minus_15)


class BudgetYearCreatePrefillTests(TestCase):
    """Test that creating a new budget year pre-fills sub-budgets from previous year."""

    def setUp(self):
        self.house = House.objects.create(code="BB", name="Maison BB", account_number="13-51200")
        self.prev_by = BudgetYear.objects.create(
            house=self.house, year=2025,
            annual_budget_total=Decimal("10000.00"),
        )
        # Add sub-budgets to previous year
        SubBudget.objects.create(
            budget_year=self.prev_by, trace_code=1,
            name="Réparations apt", planned_amount=Decimal("2000.00"), sort_order=1,
        )
        SubBudget.objects.create(
            budget_year=self.prev_by, trace_code=7,
            name="Produits ménager", planned_amount=Decimal("500.00"), sort_order=7,
        )

    def test_previous_sub_budgets_returned(self):
        """_get_previous_sub_budgets should return previous year's active non-contingency subs."""
        view = BudgetYearCreateView()
        view.request = RequestFactory().get("/budget/create/")
        subs = view._get_previous_sub_budgets(self.house)
        names = [s["name"] for s in subs]
        self.assertIn("Réparations apt", names)
        self.assertIn("Produits ménager", names)
        # Contingency (trace_code=0) should NOT be included
        codes = [s["trace_code"] for s in subs]
        self.assertNotIn(0, codes)

    def test_seed_defaults_when_no_previous_year(self):
        """Without a previous year, seed defaults should be returned."""
        other_house = House.objects.create(code="CC", name="Maison CC", account_number="13-51300")
        view = BudgetYearCreateView()
        view.request = RequestFactory().get("/budget/create/")
        subs = view._get_previous_sub_budgets(other_house)
        # Should have SEED_CATEGORIES (13 items)
        self.assertEqual(len(subs), 13)
        names = [s["name"] for s in subs]
        self.assertIn("Réparations par appartement", names)
        self.assertIn("Corvées", names)

    def test_amounts_carried_over(self):
        """Previous year's planned_amount should be carried over."""
        view = BudgetYearCreateView()
        view.request = RequestFactory().get("/budget/create/")
        subs = view._get_previous_sub_budgets(self.house)
        repairs = [s for s in subs if s["trace_code"] == 1][0]
        self.assertEqual(repairs["planned_amount"], Decimal("2000.00"))


class BudgetYearListFilteringTests(TestCase):
    def setUp(self):
        self.house_bb = House.objects.create(code="BB", name="Maison BB", account_number="13-51200")
        self.house_db = House.objects.create(code="DB", name="Maison DB", account_number="16-51200")
        self.bb_2026 = BudgetYear.objects.create(
            house=self.house_bb, year=2026, annual_budget_total=Decimal("12237.00")
        )
        self.bb_2025 = BudgetYear.objects.create(
            house=self.house_bb, year=2025, annual_budget_total=Decimal("12000.00")
        )
        self.db_2026 = BudgetYear.objects.create(
            house=self.house_db, year=2026, annual_budget_total=Decimal("13000.00")
        )
        self.db_2025 = BudgetYear.objects.create(
            house=self.house_db, year=2025, annual_budget_total=Decimal("12500.00")
        )
        self.user = User.objects.create_user(
            username="tresorier-bb",
            password="test123",
            role=User.Role.TREASURER,
            house=self.house_bb,
        )

    def test_year_list_defaults_to_current_users_house(self):
        self.client.login(username="tresorier-bb", password="test123")
        response = self.client.get(reverse("budget:year-list"))
        self.assertEqual(response.status_code, 200)
        # Defaults to current year (2026) and user's house (BB)
        current_year = date.today().year
        self.assertEqual(
            [(by.house.code, by.year) for by in response.context["budget_years"]],
            [("BB", current_year)],
        )
        self.assertEqual(response.context["selected_house"], str(self.house_bb.pk))

    def test_year_list_can_filter_by_other_house(self):
        self.client.login(username="tresorier-bb", password="test123")
        response = self.client.get(
            reverse("budget:year-list"),
            {"house": str(self.house_db.pk), "year": ""},
        )
        self.assertEqual(
            [(by.house.code, by.year) for by in response.context["budget_years"]],
            [("DB", 2026), ("DB", 2025)],
        )

    def test_year_list_can_filter_by_year_across_all_houses(self):
        self.client.login(username="tresorier-bb", password="test123")
        response = self.client.get(
            reverse("budget:year-list"),
            {"house": "", "year": "2025"},
        )
        self.assertEqual(
            [(by.house.code, by.year) for by in response.context["budget_years"]],
            [("BB", 2025), ("DB", 2025)],
        )

    def test_year_list_orders_years_descending(self):
        self.client.login(username="tresorier-bb", password="test123")
        response = self.client.get(reverse("budget:year-list"), {"house": "", "year": ""})
        years = [by.year for by in response.context["budget_years"]]
        self.assertEqual(years, [2026, 2026, 2025, 2025])

    def test_year_detail_is_read_only_for_other_house_treasurer(self):
        self.client.login(username="tresorier-bb", password="test123")
        response = self.client.get(reverse("budget:year-detail", kwargs={"pk": self.db_2026.pk}))
        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, "Ajouter une dépense")
        self.assertNotContains(response, "Modifier le budget")

    def test_expense_ledger_is_visible_for_other_house(self):
        self.client.login(username="tresorier-bb", password="test123")
        response = self.client.get(
            reverse("budget:expense-ledger", kwargs={"budget_year_pk": self.db_2026.pk})
        )
        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, "Ajouter une dépense")


class BudgetLedgerSignerDisplayTests(TestCase):
    def setUp(self):
        self.house = House.objects.create(code="BB", name="Maison BB", account_number="13-51200")
        self.by = BudgetYear.objects.create(
            house=self.house, year=2026, annual_budget_total=Decimal("12237.00")
        )
        self.sub_budget = SubBudget.objects.create(
            budget_year=self.by, trace_code=7,
            name="Produits ménager", planned_amount=Decimal("300.00")
        )
        self.purchaser = Member.objects.create(first_name="Marylin", last_name="Lamarche")
        self.approver = Member.objects.create(first_name="René", last_name="Côté")
        self.treasurer_member = Member.objects.create(first_name="Trésorier", last_name="Test")
        self.apt_202 = Apartment.objects.create(house=self.house, code="202")
        self.apt_203 = Apartment.objects.create(house=self.house, code="203")
        self.apt_204 = Apartment.objects.create(house=self.house, code="204")
        Residency.objects.create(member=self.purchaser, apartment=self.apt_202, start_date=date(2020, 1, 1))
        Residency.objects.create(member=self.approver, apartment=self.apt_203, start_date=date(2020, 1, 1))
        Residency.objects.create(member=self.treasurer_member, apartment=self.apt_204, start_date=date(2020, 1, 1))
        self.user = User.objects.create_user(
            username="tresorier",
            password="test123",
            role=User.Role.TREASURER,
            house=self.house,
            member=self.treasurer_member,
        )

    def test_year_detail_shows_validator_column_with_treasurer_fallback(self):
        bon = BonDeCommande.objects.create(
            house=self.house,
            budget_year=self.by,
            number="16011",
            purchase_date=date(2026, 1, 7),
            short_description="Paper BC",
            total=Decimal("25.00"),
            sub_budget=self.sub_budget,
            purchaser_member=self.purchaser,
            purchaser_apartment=self.apt_202,
            status=BonStatus.VALIDATED,
            validated_by=self.user,
        )
        Expense.objects.create(
            budget_year=self.by,
            sub_budget=self.sub_budget,
            bon_de_commande=bon,
            entry_date=date(2026, 1, 7),
            description="Nettoyants",
            bon_number=bon.number,
            supplier_name="RONA",
            spent_by_label="stale",
            amount=Decimal("25.00"),
        )

        self.client.login(username="tresorier", password="test123")
        response = self.client.get(reverse("budget:year-detail", kwargs={"pk": self.by.pk}))
        self.assertContains(response, "Validé par")
        self.assertContains(response, "202 / Marylin Lamarche")
        self.assertContains(response, "204 / Trésorier Test")

    def test_expense_ledger_prefers_explicit_bon_approver_label(self):
        bon = BonDeCommande.objects.create(
            house=self.house,
            budget_year=self.by,
            number="16012",
            purchase_date=date(2026, 1, 8),
            short_description="Paper BC",
            total=Decimal("50.00"),
            sub_budget=self.sub_budget,
            purchaser_member=self.purchaser,
            purchaser_apartment=self.apt_202,
            approver_member=self.approver,
            approver_apartment=self.apt_203,
            status=BonStatus.VALIDATED,
            validated_by=self.user,
        )
        Expense.objects.create(
            budget_year=self.by,
            sub_budget=self.sub_budget,
            bon_de_commande=bon,
            entry_date=date(2026, 1, 8),
            description="Peinture",
            bon_number=bon.number,
            supplier_name="RONA",
            spent_by_label="stale",
            amount=Decimal("50.00"),
        )

        self.client.login(username="tresorier", password="test123")
        response = self.client.get(reverse("budget:expense-ledger", kwargs={"budget_year_pk": self.by.pk}))
        self.assertContains(response, "Validé par")
        self.assertContains(response, "202 / Marylin Lamarche")
        self.assertContains(response, "203 / René Côté")


class ExpenseDateEditTests(TestCase):
    """Test that the HTML5 date input is populated when editing an expense."""

    def setUp(self):
        self.house = House.objects.create(code="BB", name="Maison BB", account_number="13-51200")
        self.by = BudgetYear.objects.create(
            house=self.house, year=2026, annual_budget_total=Decimal("12237.00")
        )
        self.sub = SubBudget.objects.create(
            budget_year=self.by, trace_code=7,
            name="Produits ménager", planned_amount=Decimal("300.00")
        )
        self.user = User.objects.create_user(
            username="tresorier", password="test123",
            role=User.Role.TREASURER, house=self.house,
        )
        self.expense = Expense.objects.create(
            budget_year=self.by, sub_budget=self.sub,
            entry_date=date(2026, 3, 15), description="Nettoyants",
            amount=Decimal("19.49"), spent_by_label="202 / Marylin"
        )

    def test_edit_form_shows_date_value(self):
        self.client.login(username="tresorier", password="test123")
        response = self.client.get(reverse("budget:expense-edit", kwargs={"pk": self.expense.pk}))
        self.assertContains(response, 'value="2026-03-15"')


class ExpenseCancellationTests(TestCase):
    """Test expense cancellation creates a reversal entry."""

    def setUp(self):
        self.house = House.objects.create(code="BB", name="Maison BB", account_number="13-51200")
        self.by = BudgetYear.objects.create(
            house=self.house, year=2026, annual_budget_total=Decimal("12237.00")
        )
        self.sub = SubBudget.objects.create(
            budget_year=self.by, trace_code=7,
            name="Produits ménager", planned_amount=Decimal("300.00")
        )
        self.user = User.objects.create_user(
            username="tresorier", password="test123",
            role=User.Role.TREASURER, house=self.house,
        )
        self.expense = Expense.objects.create(
            budget_year=self.by, sub_budget=self.sub,
            entry_date=date(2026, 3, 1), description="Nettoyants",
            amount=Decimal("19.49"), spent_by_label="202 / Marylin",
            entered_by=self.user,
        )

    def test_cancel_creates_reversal(self):
        self.client.login(username="tresorier", password="test123")
        response = self.client.post(
            reverse("budget:expense-cancel", kwargs={"pk": self.expense.pk})
        )
        self.assertEqual(response.status_code, 302)
        reversal = Expense.objects.filter(is_cancellation=True).first()
        self.assertIsNotNone(reversal)
        self.assertEqual(reversal.amount, Decimal("-19.49"))
        self.assertEqual(reversal.reversal_of, self.expense)
        self.assertIn("ANNULATION", reversal.description)
        self.assertEqual(reversal.sub_budget, self.sub)

    def test_cancel_already_cancelled_is_rejected(self):
        self.client.login(username="tresorier", password="test123")
        self.client.post(reverse("budget:expense-cancel", kwargs={"pk": self.expense.pk}))
        response = self.client.post(
            reverse("budget:expense-cancel", kwargs={"pk": self.expense.pk})
        )
        self.assertEqual(response.status_code, 302)
        # Should still be just one reversal
        self.assertEqual(Expense.objects.filter(is_cancellation=True).count(), 1)

    def test_cancel_cancellation_is_rejected(self):
        self.client.login(username="tresorier", password="test123")
        self.client.post(reverse("budget:expense-cancel", kwargs={"pk": self.expense.pk}))
        reversal = Expense.objects.get(is_cancellation=True)
        response = self.client.post(
            reverse("budget:expense-cancel", kwargs={"pk": reversal.pk})
        )
        self.assertEqual(response.status_code, 302)
        # Still just one reversal
        self.assertEqual(Expense.objects.filter(is_cancellation=True).count(), 1)

    def test_running_balances_include_reversal(self):
        """Running balance should reflect the cancellation (net zero)."""
        self.client.login(username="tresorier", password="test123")
        self.client.post(reverse("budget:expense-cancel", kwargs={"pk": self.expense.pk}))
        balances = BudgetCalculationService.running_balances(self.by)
        self.assertEqual(len(balances), 2)
        # After original: 12237 - 19.49 = 12217.51
        # After reversal: 12217.51 - (-19.49) = 12237.00
        self.assertEqual(balances[1]["balance"], Decimal("12237.00"))

    def test_running_balances_can_hide_cancelled_entries(self):
        self.client.login(username="tresorier", password="test123")
        self.client.post(reverse("budget:expense-cancel", kwargs={"pk": self.expense.pk}), follow=True)
        balances = BudgetCalculationService.running_balances(
            self.by,
            include_cancelled=False,
        )
        self.assertEqual(balances, [])

    def test_ledger_shows_cancellation_in_red(self):
        self.client.login(username="tresorier", password="test123")
        self.client.post(reverse("budget:expense-cancel", kwargs={"pk": self.expense.pk}))
        response = self.client.get(
            reverse("budget:expense-ledger", kwargs={"budget_year_pk": self.by.pk}) + "?show_cancelled=1"
        )
        self.assertContains(response, "[ANNULATION]")
        self.assertContains(response, "var(--negative, #c0392b)")
        self.assertContains(response, "-19.49")

    def test_ledger_hides_cancelled_entries_by_default(self):
        self.client.login(username="tresorier", password="test123")
        self.client.post(reverse("budget:expense-cancel", kwargs={"pk": self.expense.pk}), follow=True)
        response = self.client.get(
            reverse("budget:expense-ledger", kwargs={"budget_year_pk": self.by.pk})
        )
        self.assertNotContains(response, "Nettoyants")
        self.assertNotContains(response, "[ANNULATION]")
        self.assertContains(response, "Afficher les dépenses annulées")

    def test_year_detail_hides_cancelled_entries_by_default(self):
        self.client.login(username="tresorier", password="test123")
        self.client.post(reverse("budget:expense-cancel", kwargs={"pk": self.expense.pk}), follow=True)
        response = self.client.get(
            reverse("budget:year-detail", kwargs={"pk": self.by.pk})
        )
        self.assertNotContains(response, "Nettoyants")
        self.assertNotContains(response, "[ANNULATION]")

    def test_show_cancelled_query_parameter_restores_entries(self):
        self.client.login(username="tresorier", password="test123")
        self.client.post(reverse("budget:expense-cancel", kwargs={"pk": self.expense.pk}), follow=True)
        response = self.client.get(
            reverse("budget:year-detail", kwargs={"pk": self.by.pk}) + "?show_cancelled=1"
        )
        self.assertContains(response, "Nettoyants")
        self.assertContains(response, "[ANNULATION]")

    def test_edit_form_shows_cancel_button(self):
        self.client.login(username="tresorier", password="test123")
        response = self.client.get(reverse("budget:expense-edit", kwargs={"pk": self.expense.pk}))
        self.assertContains(response, "Annuler cette dépense")

    def test_edit_form_hides_cancel_for_already_cancelled(self):
        self.client.login(username="tresorier", password="test123")
        self.client.post(reverse("budget:expense-cancel", kwargs={"pk": self.expense.pk}))
        response = self.client.get(reverse("budget:expense-edit", kwargs={"pk": self.expense.pk}))
        self.assertNotContains(response, "Annuler cette dépense")

    def test_edit_form_shows_reactivate_button_for_cancelled_expense(self):
        self.client.login(username="tresorier", password="test123")
        self.client.post(reverse("budget:expense-cancel", kwargs={"pk": self.expense.pk}), follow=True)
        response = self.client.get(reverse("budget:expense-edit", kwargs={"pk": self.expense.pk}))
        self.assertContains(response, "Réactiver cette dépense")

    def test_reactivate_removes_reversal(self):
        self.client.login(username="tresorier", password="test123")
        self.client.post(reverse("budget:expense-cancel", kwargs={"pk": self.expense.pk}))
        response = self.client.post(
            reverse("budget:expense-reactivate", kwargs={"pk": self.expense.pk})
        )
        self.assertEqual(response.status_code, 302)
        self.assertFalse(self.expense.reversals.exists())
        self.assertFalse(Expense.objects.filter(is_cancellation=True).exists())

    def test_model_rejects_negative_for_non_cancellation(self):
        """Non-cancellation expenses cannot have negative amounts."""
        from django.core.exceptions import ValidationError
        exp = Expense(
            budget_year=self.by, sub_budget=self.sub,
            entry_date=date(2026, 3, 2), description="Invalid",
            amount=Decimal("-10.00"), spent_by_label="test",
        )
        with self.assertRaises(ValidationError):
            exp.full_clean()


class ExpenseReceiptPageTests(TestCase):
    def setUp(self):
        self.house = House.objects.create(code="BB", name="Maison BB", account_number="13-51200")
        self.by = BudgetYear.objects.create(
            house=self.house, year=2026, annual_budget_total=Decimal("12237.00")
        )
        self.sub = SubBudget.objects.create(
            budget_year=self.by, trace_code=7,
            name="Produits ménager", planned_amount=Decimal("300.00")
        )
        self.user = User.objects.create_user(
            username="tresorier", password="test123",
            role=User.Role.TREASURER, house=self.house,
        )
        self.purchaser = Member.objects.create(first_name="Marylin", last_name="Lamarche")
        self.apartment = Apartment.objects.create(house=self.house, code="202")
        Residency.objects.create(
            member=self.purchaser,
            apartment=self.apartment,
            start_date=date(2020, 1, 1),
        )
        self.bon = BonDeCommande.objects.create(
            house=self.house,
            budget_year=self.by,
            number="16099",
            purchase_date=date(2026, 3, 20),
            short_description="Produits",
            total=Decimal("19.49"),
            sub_budget=self.sub,
            purchaser_member=self.purchaser,
            purchaser_apartment=self.apartment,
            status=BonStatus.VALIDATED,
            validated_by=self.user,
        )
        self.expense = Expense.objects.create(
            budget_year=self.by,
            sub_budget=self.sub,
            bon_de_commande=self.bon,
            entry_date=date(2026, 3, 20),
            description="Nettoyants",
            bon_number=self.bon.number,
            supplier_name="RONA",
            spent_by_label="202 / Marylin",
            amount=Decimal("19.49"),
        )
        self.receipt = ReceiptFile.objects.create(
            bon_de_commande=self.bon,
            file=SimpleUploadedFile("receipt1.png", b"fake image data", content_type="image/png"),
            original_filename="receipt1.png",
            content_type="image/png",
            uploaded_by=self.user,
            ocr_status=OcrStatus.EXTRACTED,
        )

    def test_expense_grid_shows_receipt_link(self):
        self.client.login(username="tresorier", password="test123")
        response = self.client.get(reverse("budget:expense-ledger", kwargs={"budget_year_pk": self.by.pk}))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, reverse("budget:expense-receipts", kwargs={"pk": self.expense.pk}))
        self.assertContains(response, "Voir (1)")

    def test_year_detail_shows_receipt_link(self):
        self.client.login(username="tresorier", password="test123")
        response = self.client.get(reverse("budget:year-detail", kwargs={"pk": self.by.pk}))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Factures")
        self.assertContains(response, reverse("budget:expense-receipts", kwargs={"pk": self.expense.pk}))

    def test_receipt_page_lists_active_receipts(self):
        self.client.login(username="tresorier", password="test123")
        response = self.client.get(reverse("budget:expense-receipts", kwargs={"pk": self.expense.pk}))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "receipt1.png")
        self.assertContains(response, reverse("bons:detail", kwargs={"pk": self.bon.pk}))

    def test_receipt_page_handles_expense_without_bon(self):
        standalone = Expense.objects.create(
            budget_year=self.by,
            sub_budget=self.sub,
            entry_date=date(2026, 3, 21),
            description="Dépense directe",
            supplier_name="Ville",
            spent_by_label="BB",
            amount=Decimal("42.00"),
            source_type="accountant_direct",
        )
        self.client.login(username="tresorier", password="test123")
        response = self.client.get(reverse("budget:expense-receipts", kwargs={"pk": standalone.pk}))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Aucune facture active associée à cette dépense.")


# ──────────────────────────────────────────────────────────────────
#  Grille export (PDF / XLSX) tests
# ──────────────────────────────────────────────────────────────────

class ExpenseLedgerExportTests(TestCase):
    """Test PDF and Excel export of the expense ledger."""

    def setUp(self):
        self.house = House.objects.create(code="BB", name="Maison BB", account_number="13-51200")
        self.by = BudgetYear.objects.create(
            house=self.house, year=2025,
            annual_budget_total=Decimal("11951.00"),
            snow_budget=Decimal("1858.00"),
        )
        self.sub = SubBudget.objects.create(
            budget_year=self.by, trace_code=1,
            name="Réparations", planned_amount=Decimal("2080.00"),
        )
        Expense.objects.create(
            budget_year=self.by, sub_budget=self.sub,
            entry_date=date(2025, 3, 15), description="Plomberie",
            bon_number="BB250001", supplier_name="Plombier Joe",
            spent_by_label="M. Test",
            amount=Decimal("125.50"),
        )
        Expense.objects.create(
            budget_year=self.by, sub_budget=self.sub,
            entry_date=date(2025, 4, 1), description="Peinture",
            bon_number="BB250002", supplier_name="Rona",
            spent_by_label="M. Test",
            amount=Decimal("78.22"),
        )
        self.user = User.objects.create_user(
            username="tresorier", password="test123",
            role=User.Role.TREASURER, house=self.house,
        )
        self.client.login(username="tresorier", password="test123")

    def test_pdf_export_returns_pdf(self):
        url = reverse("budget:expense-ledger-pdf", kwargs={"budget_year_pk": self.by.pk})
        resp = self.client.get(url)
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp["Content-Type"], "application/pdf")
        self.assertIn("Grille_depenses_BB_2025.pdf", resp["Content-Disposition"])
        self.assertTrue(resp.content[:5] == b"%PDF-")

    def test_xlsx_export_returns_xlsx(self):
        url = reverse("budget:expense-ledger-xlsx", kwargs={"budget_year_pk": self.by.pk})
        resp = self.client.get(url)
        self.assertEqual(resp.status_code, 200)
        self.assertIn("spreadsheetml", resp["Content-Type"])
        self.assertIn("Grille_depenses_BB_2025.xlsx", resp["Content-Disposition"])
        # Verify it's a valid XLSX (ZIP magic bytes)
        self.assertTrue(resp.content[:2] == b"PK")

    def test_xlsx_contains_expense_data(self):
        """Verify the XLSX contains our expense rows."""
        import openpyxl
        import io
        url = reverse("budget:expense-ledger-xlsx", kwargs={"budget_year_pk": self.by.pk})
        resp = self.client.get(url)
        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        ws = wb["Grille de dépenses"]
        # Header row is 4; data starts at 5
        self.assertEqual(ws.cell(row=5, column=2).value, "Plomberie")
        self.assertEqual(ws.cell(row=6, column=2).value, "Peinture")
        self.assertAlmostEqual(ws.cell(row=5, column=9).value, 125.50)
        self.assertAlmostEqual(ws.cell(row=6, column=9).value, 78.22)
        # Summary sheet exists
        self.assertIn("Résumé budgétaire", wb.sheetnames)

    def test_pdf_export_with_cancelled_param(self):
        url = reverse("budget:expense-ledger-pdf", kwargs={"budget_year_pk": self.by.pk})
        resp = self.client.get(url + "?show_cancelled=1")
        self.assertEqual(resp.status_code, 200)

    def test_export_requires_login(self):
        self.client.logout()
        url = reverse("budget:expense-ledger-pdf", kwargs={"budget_year_pk": self.by.pk})
        resp = self.client.get(url)
        self.assertNotEqual(resp.status_code, 200)


# ---------------------------------------------------------------------------
# Budget Year Inactive Protection Tests
# ---------------------------------------------------------------------------

class BudgetYearInactiveProtectionTests(TestCase):
    """Verify that past-year budgets block expense mutations but allow GL import."""

    def setUp(self):
        self.house = House.objects.create(code="BB", name="Maison BB", account_number="13-51200")
        current_year = date.today().year

        # Past-year budget (inactive by calendar)
        self.past_by = BudgetYear.objects.create(
            house=self.house, year=current_year - 1,
            annual_budget_total=Decimal("10000.00"),
        )
        self.past_sub = SubBudget.objects.create(
            budget_year=self.past_by, trace_code=7,
            name="Produits ménager", planned_amount=Decimal("300.00"),
        )

        # Current-year budget (active)
        self.current_by = BudgetYear.objects.create(
            house=self.house, year=current_year,
            annual_budget_total=Decimal("12000.00"),
        )
        self.current_sub = SubBudget.objects.create(
            budget_year=self.current_by, trace_code=7,
            name="Produits ménager", planned_amount=Decimal("300.00"),
        )

        self.user = User.objects.create_user(
            username="tresorier", password="test123",
            role=User.Role.TREASURER, house=self.house,
        )

        self.past_expense = Expense.objects.create(
            budget_year=self.past_by, sub_budget=self.past_sub,
            entry_date=date(current_year - 1, 6, 1),
            description="Old expense", amount=Decimal("50.00"),
            spent_by_label="202 / Marylin", entered_by=self.user,
        )

    # -- Model property tests ------------------------------------------------

    def test_is_active_property_current_year(self):
        """BudgetYear for the current year is active."""
        self.assertTrue(self.current_by.is_year_active)
        self.assertFalse(self.current_by.is_inactive)

    def test_is_active_property_past_year(self):
        """BudgetYear for a past year is inactive."""
        self.assertFalse(self.past_by.is_year_active)
        self.assertTrue(self.past_by.is_inactive)

    # -- Expense creation blocked on inactive budget --------------------------

    def test_cannot_create_expense_on_inactive_budget(self):
        self.client.login(username="tresorier", password="test123")
        url = reverse("budget:expense-create", kwargs={"budget_year_pk": self.past_by.pk})
        resp = self.client.post(url, {
            "sub_budget": self.past_sub.pk,
            "entry_date": "2024-06-01",
            "description": "Should be blocked",
            "amount": "25.00",
            "spent_by_label": "202 / Marylin",
        })
        self.assertEqual(resp.status_code, 403)

    def test_cannot_create_expense_get_on_inactive_budget(self):
        self.client.login(username="tresorier", password="test123")
        url = reverse("budget:expense-create", kwargs={"budget_year_pk": self.past_by.pk})
        resp = self.client.get(url)
        self.assertEqual(resp.status_code, 403)

    # -- Expense edit blocked on inactive budget ------------------------------

    def test_cannot_edit_expense_on_inactive_budget(self):
        self.client.login(username="tresorier", password="test123")
        url = reverse("budget:expense-edit", kwargs={"pk": self.past_expense.pk})
        resp = self.client.get(url)
        self.assertEqual(resp.status_code, 403)

    # -- Expense cancel / reactivate blocked on inactive budget ---------------

    def test_cannot_cancel_expense_on_inactive_budget(self):
        self.client.login(username="tresorier", password="test123")
        url = reverse("budget:expense-cancel", kwargs={"pk": self.past_expense.pk})
        resp = self.client.post(url)
        self.assertEqual(resp.status_code, 403)

    def test_cannot_reactivate_expense_on_inactive_budget(self):
        """Even if a reversal exists, reactivation is blocked on inactive budget."""
        # Create a reversal manually for the past expense
        Expense.objects.create(
            budget_year=self.past_by, sub_budget=self.past_sub,
            entry_date=date.today(), description="[ANNULATION] Old expense",
            amount=Decimal("-50.00"), spent_by_label="202 / Marylin",
            entered_by=self.user, is_cancellation=True,
            reversal_of=self.past_expense,
        )
        self.client.login(username="tresorier", password="test123")
        url = reverse("budget:expense-reactivate", kwargs={"pk": self.past_expense.pk})
        resp = self.client.post(url)
        self.assertEqual(resp.status_code, 403)

    # -- GL import still allowed on inactive budget ---------------------------

    def test_can_import_gl_entries_on_inactive_budget(self):
        """GrandLivreValidateView must NOT be blocked for past-year budgets."""
        from budget.models import GrandLivreUpload, GrandLivreEntry

        upload = GrandLivreUpload.objects.create(
            budget_year=self.past_by, uploaded_by=self.user,
            status="parsed",
        )
        entry = GrandLivreEntry.objects.create(
            upload=upload, row_number=1,
            date=date(date.today().year - 1, 12, 15),
            description_raw="Plomberie urgence",
            description_clean="Plomberie urgence",
            source="Fournisseur ABC",
            debit=Decimal("150.00"), credit=Decimal("0.00"),
            match_confidence="unmatched",
            needs_import=True,
        )

        self.client.login(username="tresorier", password="test123")
        url = reverse("budget:grand-livre-validate", kwargs={"pk": upload.pk})
        resp = self.client.post(url, {"entry_ids": [str(entry.pk)]})
        # Should redirect (302), NOT 403
        self.assertIn(resp.status_code, [200, 302])
        self.assertNotEqual(resp.status_code, 403)

    # -- Active budget still works normally -----------------------------------

    def test_can_create_expense_on_active_budget(self):
        self.client.login(username="tresorier", password="test123")
        url = reverse("budget:expense-create", kwargs={"budget_year_pk": self.current_by.pk})
        resp = self.client.get(url)
        # GET should render the form (200), not 403
        self.assertEqual(resp.status_code, 200)
