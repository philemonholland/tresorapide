"""
Grand Livre Excel parser.

Reads the accountant's Grand Livre Excel file, locates the section
for a given house account number, and extracts all transaction rows.
"""
import re
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import List, Optional

import openpyxl


@dataclass
class GLTransaction:
    """A single parsed transaction from the Grand Livre."""
    row_number: int
    period: str = ""
    date: Optional[date] = None
    source: str = ""
    description: str = ""
    debit: Decimal = Decimal("0")
    credit: Decimal = Decimal("0")
    solde_fin: Optional[Decimal] = None


@dataclass
class GLAccountSection:
    """Parsed section for a single account from the Grand Livre."""
    account_number: str = ""
    description: str = ""
    period_end_date: Optional[date] = None
    transactions: List[GLTransaction] = field(default_factory=list)
    total_debit: Decimal = Decimal("0")
    total_credit: Decimal = Decimal("0")
    solde_fin: Decimal = Decimal("0")
    entry_count: int = 0


HEADER_KEYWORDS = {"No compte", "Description", "Solde fin", "Débit", "Crédit"}

# French month names → month numbers
_FRENCH_MONTHS = {
    "janvier": 1, "février": 2, "mars": 3, "avril": 4, "mai": 5, "juin": 6,
    "juillet": 7, "août": 8, "septembre": 9, "octobre": 10, "novembre": 11,
    "décembre": 12,
}
_FR_DATE_PATTERN = re.compile(
    r"(\d{1,2})\s+(janvier|février|mars|avril|mai|juin|juillet|août|"
    r"septembre|octobre|novembre|décembre)\s+(\d{4})",
    re.IGNORECASE,
)


def _parse_french_date(text: str) -> Optional[date]:
    """Parse a French date like '03 mars 2026' from a text string."""
    m = _FR_DATE_PATTERN.search(text)
    if m:
        day = int(m.group(1))
        month = _FRENCH_MONTHS.get(m.group(2).lower())
        year = int(m.group(3))
        if month:
            try:
                return date(year, month, day)
            except ValueError:
                pass
    return None


def _extract_period_end_date(ws) -> Optional[date]:
    """Scan the header area (rows 1–8) for the GL period end date.

    The date typically appears in a cell like
    ``"Total des dépenses au 03 mars 2026"`` (row 5).
    We also check for a datetime value in column D (4) or K (11).
    """
    for row_idx in range(1, 9):
        for col_idx in range(1, 12):
            val = ws.cell(row_idx, col_idx).value
            if val is None:
                continue
            # Try French text
            if isinstance(val, str) and "au " in val.lower():
                dt = _parse_french_date(val)
                if dt:
                    return dt
            # Try direct datetime
            if isinstance(val, datetime):
                return val.date()
            if isinstance(val, date) and not isinstance(val, datetime):
                return val
    return None


def _to_decimal(value) -> Decimal:
    if value is None:
        return Decimal("0")
    try:
        return Decimal(str(value)).quantize(Decimal("0.01"))
    except (InvalidOperation, ValueError):
        return Decimal("0")


def _to_date(value) -> Optional[date]:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def _normalize_account_number(raw: str) -> str:
    """Normalize account number: '13-51200' or '1351200' → '13-51200'."""
    raw = raw.strip().replace(" ", "")
    if "-" in raw:
        return raw
    match = re.match(r"^(\d{1,3})(51200)$", raw)
    if match:
        return f"{match.group(1)}-{match.group(2)}"
    return raw


def _find_header_row(ws) -> int:
    """Find the row containing column headers."""
    for row_idx in range(1, min(20, ws.max_row + 1)):
        vals = {str(c.value or "").strip() for c in ws[row_idx]}
        if len(vals & HEADER_KEYWORDS) >= 3:
            return row_idx
    return 8  # fallback


def parse_grand_livre(file_path, target_account: str) -> GLAccountSection:
    """
    Parse the Grand Livre Excel file and extract the section for
    the given house account number (e.g. '13-51200').

    Returns a GLAccountSection with all transactions for that account.
    """
    wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    ws = wb.active or wb[wb.sheetnames[0]]
    target_normalized = _normalize_account_number(target_account)
    target_prefix = target_normalized.split("-")[0] if "-" in target_normalized else target_normalized

    header_row = _find_header_row(ws)
    result = GLAccountSection(account_number=target_normalized)
    result.period_end_date = _extract_period_end_date(ws)
    in_target_section = False
    current_period = ""

    for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=False), header_row + 1):
        vals = [cell.value for cell in row]
        a_val = vals[0]  # No compte
        b_val = vals[1] if len(vals) > 1 else None  # Description (account)
        c_val = vals[2] if len(vals) > 2 else None  # Ann/Pér
        d_val = vals[3] if len(vals) > 3 else None  # Date
        e_val = vals[4] if len(vals) > 4 else None  # Source
        f_val = vals[5] if len(vals) > 5 else None  # Description (transaction)
        g_val = vals[6] if len(vals) > 6 else None  # Total label column
        # h_val = vals[7]  # Solde début
        i_val = vals[8] if len(vals) > 8 else None  # Débit
        j_val = vals[9] if len(vals) > 9 else None  # Crédit
        k_val = vals[10] if len(vals) > 10 else None  # Solde fin

        # Check if this is an account header row. In real GL files, the first
        # transaction can live on the same row as the account header.
        if a_val and isinstance(a_val, str) and "51200" in a_val:
            acct_norm = _normalize_account_number(a_val)
            if acct_norm == target_normalized:
                in_target_section = True
                if b_val:
                    result.description = str(b_val).strip()
                has_inline_transaction = any(
                    val not in (None, "")
                    for val in (c_val, d_val, e_val, f_val, i_val, j_val, k_val)
                )
                if not has_inline_transaction:
                    continue
            elif in_target_section:
                # We've left the target section and hit a new account
                break
            else:
                continue

        if not in_target_section:
            # Check for the Grand Total line to stop parsing
            if g_val and isinstance(g_val, str) and "Grand Total" in g_val:
                break
            continue

        # Check for the total line of our account
        if g_val and isinstance(g_val, str) and "Total" in g_val:
            # Parse: "Total No compte 1351200 : 44"
            total_num = target_prefix + "51200"
            if total_num in str(g_val).replace(" ", "").replace("-", ""):
                result.total_debit = _to_decimal(i_val)
                result.total_credit = _to_decimal(j_val)
                result.solde_fin = _to_decimal(k_val)
                match = re.search(r":\s*(\d+)", str(g_val))
                if match:
                    result.entry_count = int(match.group(1))
                # Don't break — there might be a second section for same account
                in_target_section = False
                continue

        # Track period changes
        if c_val:
            dt = _to_date(c_val)
            if dt:
                current_period = dt.strftime("%Y-%m")
            else:
                current_period = str(c_val).strip()

        # Skip empty rows
        if not d_val and not f_val and not i_val:
            continue

        # Parse transaction row
        tx = GLTransaction(
            row_number=row_idx,
            period=current_period,
            date=_to_date(d_val),
            source=str(e_val).strip() if e_val else "",
            description=str(f_val).strip() if f_val else "",
            debit=_to_decimal(i_val),
            credit=_to_decimal(j_val),
            solde_fin=_to_decimal(k_val) if k_val else None,
        )
        result.transactions.append(tx)

    wb.close()

    # If we found entries across multiple sections (same account split),
    # aggregate the totals
    if not result.transactions:
        return result

    # Recalculate entry count from actual transactions if the parsed count
    # didn't match (can happen with multi-section accounts)
    if result.entry_count == 0:
        result.entry_count = len(result.transactions)

    return result


def parse_all_accounts(file_path) -> dict:
    """
    Parse all account sections in the Grand Livre.
    Returns a dict mapping account_number → GLAccountSection.
    Useful for finding which accounts exist.
    """
    wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    ws = wb.active or wb[wb.sheetnames[0]]
    header_row = _find_header_row(ws)
    accounts = {}

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        a_val = row[0] if row else None
        if a_val and isinstance(a_val, str) and "51200" in a_val:
            acct = _normalize_account_number(a_val)
            if acct not in accounts:
                accounts[acct] = {"account": acct, "description": str(row[1] or "").strip()}

    wb.close()
    return accounts
