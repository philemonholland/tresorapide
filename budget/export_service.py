"""PDF and XLSX export for the Grille de dépenses (expense ledger).

Generates landscape-oriented exports with narrow margins,
matching the on-screen layout.
"""
import io
from decimal import Decimal

from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.units import inch, cm
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Spacer, Paragraph, KeepTogether,
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.utils import get_column_letter

from .services import BudgetCalculationService


COOP_NAME = "Coopérative d'habitation des Cantons de l'Est"


def _fmt(val):
    """Format Decimal as string with 2 decimals, or empty string."""
    if val is None:
        return ""
    return f"{val:,.2f} $".replace(",", "\u00a0")


def _fmt_plain(val):
    """Format for reportlab table cells — no thousands separator issues."""
    if val is None:
        return ""
    return f"{val:.2f} $"


# ──────────────────────────────────────────────────────────────────
#  PDF export
# ──────────────────────────────────────────────────────────────────

def generate_expense_ledger_pdf(budget_year, *, include_cancelled=False):
    """Generate a landscape PDF of the expense ledger. Returns bytes."""
    svc = BudgetCalculationService
    base = svc.base_values(budget_year)
    rows = svc.running_balances(budget_year, include_cancelled=include_cancelled)
    categories = svc.category_summary(budget_year)
    repair = svc.repair_totals(budget_year)
    imprevues = svc.imprevues_totals(budget_year)
    available = svc.available_money(budget_year)

    buf = io.BytesIO()
    page = landscape(letter)
    doc = SimpleDocTemplate(
        buf, pagesize=page,
        leftMargin=0.5 * inch, rightMargin=0.5 * inch,
        topMargin=0.4 * inch, bottomMargin=0.4 * inch,
    )

    styles = getSampleStyleSheet()
    style_title = ParagraphStyle("LedgerTitle", parent=styles["Heading1"],
                                 fontSize=14, alignment=TA_CENTER, spaceAfter=4)
    style_subtitle = ParagraphStyle("LedgerSubtitle", parent=styles["Normal"],
                                    fontSize=9, alignment=TA_CENTER, spaceAfter=8)
    style_section = ParagraphStyle("SectionHead", parent=styles["Heading2"],
                                   fontSize=11, spaceBefore=12, spaceAfter=4)
    style_cell = ParagraphStyle("Cell", parent=styles["Normal"], fontSize=7,
                                leading=9)
    style_cell_r = ParagraphStyle("CellR", parent=style_cell, alignment=TA_RIGHT)
    style_cell_c = ParagraphStyle("CellC", parent=style_cell, alignment=TA_CENTER)

    elements = []

    # Title
    house = budget_year.house
    elements.append(Paragraph(
        f"Grille de dépenses — {house.code} — {budget_year.year}", style_title
    ))
    elements.append(Paragraph(COOP_NAME, style_subtitle))
    elements.append(Spacer(1, 6))

    # ── Expense ledger table ─────────────────────────────
    header = [
        Paragraph("<b>Date</b>", style_cell),
        Paragraph("<b>Description</b>", style_cell),
        Paragraph("<b># BC</b>", style_cell),
        Paragraph("<b>GL</b>", style_cell_c),
        Paragraph("<b>Fournisseur</b>", style_cell),
        Paragraph("<b>Remb.</b>", style_cell),
        Paragraph("<b>Dépensé par</b>", style_cell),
        Paragraph("<b>Validé par</b>", style_cell),
        Paragraph("<b>Montant</b>", style_cell_r),
        Paragraph("<b>Trace</b>", style_cell_c),
        Paragraph("<b>Balance</b>", style_cell_r),
        Paragraph("<b>Balance−15%</b>", style_cell_r),
    ]

    data = [header]
    for r in rows:
        exp = r["expense"]
        desc = exp.description
        if exp.is_cancellation:
            desc += " [ANNULATION]"
        data.append([
            Paragraph(exp.entry_date.strftime("%Y-%m-%d"), style_cell),
            Paragraph(desc, style_cell),
            Paragraph(exp.bon_number or "", style_cell),
            Paragraph("✓" if exp.validated_gl else "", style_cell_c),
            Paragraph(exp.supplier_name or "", style_cell),
            Paragraph(exp.display_reimburse_label, style_cell),
            Paragraph(exp.display_spent_by_label, style_cell),
            Paragraph(exp.display_approved_by_label, style_cell),
            Paragraph(_fmt_plain(exp.amount), style_cell_r),
            Paragraph(str(exp.sub_budget.trace_code), style_cell_c),
            Paragraph(_fmt_plain(r["balance"]), style_cell_r),
            Paragraph(_fmt_plain(r["balance_minus_imprevues"]), style_cell_r),
        ])

    avail_width = page[0] - doc.leftMargin - doc.rightMargin
    col_widths = [
        0.07 * avail_width,  # Date
        0.16 * avail_width,  # Description
        0.06 * avail_width,  # BC
        0.03 * avail_width,  # GL
        0.11 * avail_width,  # Fournisseur
        0.06 * avail_width,  # Remb.
        0.11 * avail_width,  # Dépensé par
        0.11 * avail_width,  # Validé par
        0.08 * avail_width,  # Montant
        0.04 * avail_width,  # Trace
        0.08 * avail_width,  # Balance
        0.09 * avail_width,  # Balance-15%
    ]

    table = Table(data, colWidths=col_widths, repeatRows=1)
    style_cmds = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2c3e50")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("LEADING", (0, 0), (-1, -1), 9),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#bdc3c7")),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ("LEFTPADDING", (0, 0), (-1, -1), 3),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
    ]
    # Alternate row colours
    for i in range(1, len(data)):
        if i % 2 == 0:
            style_cmds.append(("BACKGROUND", (0, i), (-1, i), colors.HexColor("#f7f9fb")))
        # Red text for cancellation rows
        exp = rows[i - 1]["expense"]
        if exp.is_cancellation:
            style_cmds.append(("TEXTCOLOR", (0, i), (-1, i), colors.HexColor("#c0392b")))
        elif rows[i - 1]["is_cancelled"]:
            style_cmds.append(("TEXTCOLOR", (0, i), (-1, i), colors.HexColor("#999999")))

    table.setStyle(TableStyle(style_cmds))
    elements.append(table)

    # ── Summary table (page 2 or after ledger) ───────────
    elements.append(Spacer(1, 16))
    elements.append(Paragraph("Résumé budgétaire", style_section))

    summary_data = [
        ["Budget d'entretien de la maison", _fmt_plain(base["budget_total"])],
        ["Budget de déneigement", _fmt_plain(base["snow_budget"])],
        ["Imprévus (15 %)", _fmt_plain(base["imprevues"])],
        ["Budget − 15 %", _fmt_plain(base["budget_minus_imprevues"])],
        ["Dépenses effectuées à ce jour", _fmt_plain(base["expenses_to_date"])],
        ["Budget réparations — prévu", _fmt_plain(repair["planned"])],
        ["Budget réparations — utilisé", _fmt_plain(repair["used"])],
        ["Budget réparations — restant", _fmt_plain(repair["remaining"])],
        ["Imprévus utilisés", _fmt_plain(imprevues["used"])],
        ["Imprévus restants", _fmt_plain(imprevues["remaining"])],
        ["Argent total disponible", _fmt_plain(available["available"])],
        ["Argent total disponible − 15 %", _fmt_plain(available["available_minus_imprevues"])],
    ]
    summary_tbl = Table(summary_data, colWidths=[4 * inch, 1.5 * inch])
    summary_tbl.setStyle(TableStyle([
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("LEADING", (0, 0), (-1, -1), 11),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#bdc3c7")),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ("BACKGROUND", (0, -2), (-1, -1), colors.HexColor("#ecf0f1")),
    ]))
    elements.append(summary_tbl)

    # ── Sub-budget table ─────────────────────────────────
    if categories:
        elements.append(Spacer(1, 12))
        elements.append(Paragraph("Sous-budgets", style_section))

        cat_header = ["Description", "Trace", "Prévu", "Utilisé", "Restant"]
        cat_data = [cat_header]
        for c in categories:
            cat_data.append([
                c["name"],
                str(c["trace_code"]),
                _fmt_plain(c["planned"]),
                _fmt_plain(c["used"]),
                _fmt_plain(c["remaining"]),
            ])
        # Totals row (exclude contingency)
        non_cont = [c for c in categories if not c["sub_budget"].is_contingency]
        cat_data.append([
            "Total (excl. imprévus)",
            "",
            _fmt_plain(sum((c["planned"] for c in non_cont), Decimal("0"))),
            _fmt_plain(sum((c["used"] for c in non_cont), Decimal("0"))),
            _fmt_plain(sum((c["remaining"] for c in non_cont), Decimal("0"))),
        ])

        cat_tbl = Table(cat_data, colWidths=[3 * inch, 0.6 * inch, 1.1 * inch, 1.1 * inch, 1.1 * inch])
        cat_style = [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2c3e50")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("LEADING", (0, 0), (-1, -1), 11),
            ("ALIGN", (1, 0), (1, -1), "CENTER"),
            ("ALIGN", (2, 0), (-1, -1), "RIGHT"),
            ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#bdc3c7")),
            ("TOPPADDING", (0, 0), (-1, -1), 2),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
            ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#ecf0f1")),
            ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ]
        for i in range(1, len(cat_data) - 1):
            if i % 2 == 0:
                cat_style.append(("BACKGROUND", (0, i), (-1, i), colors.HexColor("#f7f9fb")))
        cat_tbl.setStyle(TableStyle(cat_style))
        elements.append(cat_tbl)

    doc.build(elements)
    return buf.getvalue()


# ──────────────────────────────────────────────────────────────────
#  XLSX export
# ──────────────────────────────────────────────────────────────────

_THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
_HEADER_FILL = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=9)
_ALT_FILL = PatternFill(start_color="F7F9FB", end_color="F7F9FB", fill_type="solid")
_MONEY_FMT = '#,##0.00 "$"'
_CANCEL_FONT = Font(color="C0392B", size=9)
_CANCELLED_FONT = Font(color="999999", strikethrough=True, size=9)
_TOTAL_FILL = PatternFill(start_color="ECF0F1", end_color="ECF0F1", fill_type="solid")
_BOLD_FONT = Font(bold=True, size=9)


def generate_expense_ledger_xlsx(budget_year, *, include_cancelled=False):
    """Generate landscape Excel workbook of the expense ledger. Returns bytes."""
    svc = BudgetCalculationService
    base = svc.base_values(budget_year)
    rows = svc.running_balances(budget_year, include_cancelled=include_cancelled)
    categories = svc.category_summary(budget_year)
    repair = svc.repair_totals(budget_year)
    imprevues = svc.imprevues_totals(budget_year)
    available = svc.available_money(budget_year)

    wb = openpyxl.Workbook()

    # ── Sheet 1: Grille de dépenses ──────────────────────
    ws = wb.active
    ws.title = "Grille de dépenses"
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left = 0.4
    ws.page_margins.right = 0.4
    ws.page_margins.top = 0.4
    ws.page_margins.bottom = 0.4

    house = budget_year.house
    ws.merge_cells("A1:L1")
    title_cell = ws["A1"]
    title_cell.value = f"Grille de dépenses — {house.code} — {budget_year.year}"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:L2")
    ws["A2"].value = COOP_NAME
    ws["A2"].font = Font(size=9, italic=True)
    ws["A2"].alignment = Alignment(horizontal="center")

    # Headers
    headers = [
        "Date", "Description", "# BC", "GL", "Fournisseur", "Remb.",
        "Dépensé par", "Validé par", "Montant", "Trace",
        "Balance", "Balance−15%",
    ]
    header_row = 4
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.border = _THIN_BORDER
        cell.alignment = Alignment(horizontal="center" if col_idx in (4, 10) else
                                   "right" if col_idx >= 9 else "left",
                                   wrap_text=True)

    # Data rows
    for i, r in enumerate(rows):
        row_num = header_row + 1 + i
        exp = r["expense"]
        desc = exp.description
        if exp.is_cancellation:
            desc += " [ANNULATION]"

        values = [
            exp.entry_date,
            desc,
            exp.bon_number or "",
            "✓" if exp.validated_gl else "",
            exp.supplier_name or "",
            exp.display_reimburse_label,
            exp.display_spent_by_label,
            exp.display_approved_by_label,
            float(exp.amount),
            exp.sub_budget.trace_code,
            float(r["balance"]),
            float(r["balance_minus_imprevues"]),
        ]
        for col_idx, v in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col_idx, value=v)
            cell.border = _THIN_BORDER
            cell.font = Font(size=9)
            if col_idx == 1:
                cell.number_format = "YYYY-MM-DD"
                cell.alignment = Alignment(horizontal="left")
            elif col_idx in (9, 11, 12):
                cell.number_format = _MONEY_FMT
                cell.alignment = Alignment(horizontal="right")
            elif col_idx in (4, 10):
                cell.alignment = Alignment(horizontal="center")

            # Row styling
            if exp.is_cancellation:
                cell.font = _CANCEL_FONT
            elif r["is_cancelled"]:
                cell.font = _CANCELLED_FONT

            if i % 2 == 1:
                cell.fill = _ALT_FILL

    # Column widths
    col_widths = [12, 28, 10, 5, 16, 10, 14, 14, 12, 6, 12, 12]
    for idx, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = w

    # Freeze header row
    ws.freeze_panes = f"A{header_row + 1}"

    # Print area
    last_data_row = header_row + len(rows)
    ws.print_title_rows = f"{header_row}:{header_row}"

    # ── Sheet 2: Résumé budgétaire ───────────────────────
    ws2 = wb.create_sheet("Résumé budgétaire")
    ws2.page_setup.orientation = "landscape"
    ws2.page_setup.paperSize = ws2.PAPERSIZE_LETTER
    ws2.page_margins.left = 0.5
    ws2.page_margins.right = 0.5

    ws2.merge_cells("A1:B1")
    ws2["A1"].value = "Résumé budgétaire"
    ws2["A1"].font = Font(bold=True, size=12)

    summary_items = [
        ("Budget d'entretien de la maison", base["budget_total"]),
        ("Budget de déneigement", base["snow_budget"]),
        ("Imprévus (15 %)", base["imprevues"]),
        ("Budget − 15 %", base["budget_minus_imprevues"]),
        ("Dépenses effectuées à ce jour", base["expenses_to_date"]),
        ("Budget réparations — prévu", repair["planned"]),
        ("Budget réparations — utilisé", repair["used"]),
        ("Budget réparations — restant", repair["remaining"]),
        ("Imprévus utilisés", imprevues["used"]),
        ("Imprévus restants", imprevues["remaining"]),
        ("Argent total disponible", available["available"]),
        ("Argent total disponible − 15 %", available["available_minus_imprevues"]),
    ]
    for i, (label, val) in enumerate(summary_items):
        r = 3 + i
        ws2.cell(row=r, column=1, value=label).font = Font(size=9)
        ws2.cell(row=r, column=1).border = _THIN_BORDER
        c = ws2.cell(row=r, column=2, value=float(val))
        c.number_format = _MONEY_FMT
        c.alignment = Alignment(horizontal="right")
        c.border = _THIN_BORDER
        c.font = Font(size=9)
        # Highlight last two rows
        if i >= len(summary_items) - 2:
            ws2.cell(row=r, column=1).fill = _TOTAL_FILL
            c.fill = _TOTAL_FILL
            c.font = _BOLD_FONT

    ws2.column_dimensions["A"].width = 44
    ws2.column_dimensions["B"].width = 16

    # ── Sub-budgets section on the same sheet ────────────
    sub_start = 3 + len(summary_items) + 2
    ws2.cell(row=sub_start, column=1, value="Sous-budgets").font = Font(bold=True, size=12)

    sub_headers = ["Description", "Trace", "Prévu", "Utilisé", "Restant"]
    for ci, h in enumerate(sub_headers, 1):
        cell = ws2.cell(row=sub_start + 1, column=ci, value=h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.border = _THIN_BORDER
        cell.alignment = Alignment(horizontal="right" if ci >= 3 else
                                   "center" if ci == 2 else "left")

    for i, c in enumerate(categories):
        r = sub_start + 2 + i
        ws2.cell(row=r, column=1, value=c["name"]).border = _THIN_BORDER
        ws2.cell(row=r, column=2, value=c["trace_code"]).border = _THIN_BORDER
        ws2.cell(row=r, column=2).alignment = Alignment(horizontal="center")
        for ci, key in enumerate(["planned", "used", "remaining"], 3):
            cell = ws2.cell(row=r, column=ci, value=float(c[key]))
            cell.number_format = _MONEY_FMT
            cell.alignment = Alignment(horizontal="right")
            cell.border = _THIN_BORDER
        if i % 2 == 1:
            for ci in range(1, 6):
                ws2.cell(row=r, column=ci).fill = _ALT_FILL

    # Totals row
    non_cont = [c for c in categories if not c["sub_budget"].is_contingency]
    total_row = sub_start + 2 + len(categories)
    ws2.cell(row=total_row, column=1, value="Total (excl. imprévus)").font = _BOLD_FONT
    ws2.cell(row=total_row, column=1).fill = _TOTAL_FILL
    ws2.cell(row=total_row, column=1).border = _THIN_BORDER
    ws2.cell(row=total_row, column=2).fill = _TOTAL_FILL
    ws2.cell(row=total_row, column=2).border = _THIN_BORDER
    for ci, vals in [(3, "planned"), (4, "used"), (5, "remaining")]:
        cell = ws2.cell(row=total_row, column=ci,
                        value=float(sum((c[vals] for c in non_cont), Decimal("0"))))
        cell.number_format = _MONEY_FMT
        cell.alignment = Alignment(horizontal="right")
        cell.font = _BOLD_FONT
        cell.fill = _TOTAL_FILL
        cell.border = _THIN_BORDER

    for ci in range(3, 6):
        ws2.column_dimensions[get_column_letter(ci)].width = 14

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
