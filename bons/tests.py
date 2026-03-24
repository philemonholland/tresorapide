import json
from datetime import date
from decimal import Decimal
from unittest.mock import patch, MagicMock
from django.test import TestCase
from django.core.exceptions import ValidationError
from django.core.files.uploadedfile import SimpleUploadedFile
from django.urls import reverse
from django.utils import timezone

from houses.models import House
from accounts.models import User
from members.models import Member, Apartment, Residency
from budget.models import BudgetYear, SubBudget, Expense
from bons.models import (
    BonDeCommande, BonStatus, DuplicateFlag, DuplicateFlagStatus,
    ReceiptFile, ReceiptExtractedFields, OcrStatus,
)
from bons.services import generate_bon_number
from bons.ocr_service import ReceiptOcrService, DuplicateDetectionService
from bons.scan_sessions import create_scan_session
from bons.views import (
    _names_match,
    _normalize_name,
    _bon_is_export_ready,
    _paper_bc_signer_initials,
    _resolve_member_assignment,
    _normalize_document_amounts,
)


class BonNumberGenerationTests(TestCase):
    def setUp(self):
        self.house = House.objects.create(
            code="BB", name="Maison BB", account_number="13-51200"
        )
        self.budget_year = BudgetYear.objects.create(
            house=self.house, year=2026,
            annual_budget_total=Decimal("12237.00")
        )
        self.sub_budget = SubBudget.objects.create(
            budget_year=self.budget_year, trace_code=7,
            name="Produits ménager", planned_amount=Decimal("300.00")
        )
        self.member = Member.objects.create(first_name="Marylin", last_name="Lamarche")
        self.apartment = Apartment.objects.create(house=self.house, code="202")
        Residency.objects.create(
            member=self.member, apartment=self.apartment,
            start_date=date(2020, 1, 1)
        )

    def test_first_bon_number_is_0001(self):
        number = generate_bon_number(self.house, 2026)
        self.assertEqual(number, "BB260001")

    def test_sequential_bon_numbers(self):
        BonDeCommande.objects.create(
            house=self.house, budget_year=self.budget_year,
            number="BB260001", purchase_date=date(2026, 1, 7),
            short_description="Test", total=Decimal("19.49"),
            sub_budget=self.sub_budget, purchaser_member=self.member,
        )
        number = generate_bon_number(self.house, 2026)
        self.assertEqual(number, "BB260002")

    def test_different_years_restart_sequence(self):
        BonDeCommande.objects.create(
            house=self.house, budget_year=self.budget_year,
            number="BB260005", purchase_date=date(2026, 1, 7),
            short_description="Test", total=Decimal("10.00"),
            sub_budget=self.sub_budget, purchaser_member=self.member,
        )
        number = generate_bon_number(self.house, 2027)
        self.assertEqual(number, "BB270001")

    def test_different_houses_independent_sequences(self):
        house2 = House.objects.create(code="DB", name="Maison DB", account_number="16-51200")
        number = generate_bon_number(house2, 2026)
        self.assertEqual(number, "DB260001")


class BonDeCommandeModelTests(TestCase):
    def setUp(self):
        self.house = House.objects.create(code="BB", name="Maison BB", account_number="13-51200")
        self.budget_year = BudgetYear.objects.create(
            house=self.house, year=2026, annual_budget_total=Decimal("12237.00")
        )
        self.sub_budget = SubBudget.objects.create(
            budget_year=self.budget_year, trace_code=7,
            name="Produits ménager", planned_amount=Decimal("300.00")
        )
        self.purchaser = Member.objects.create(first_name="Marylin", last_name="Lamarche")
        self.approver = Member.objects.create(first_name="Carl-David", last_name="Fortin")
        self.apt = Apartment.objects.create(house=self.house, code="202")

    def test_approver_cannot_equal_purchaser(self):
        bon = BonDeCommande(
            house=self.house, budget_year=self.budget_year, number="BB260001",
            purchase_date=date(2026, 1, 7), short_description="Test",
            total=Decimal("19.49"), sub_budget=self.sub_budget,
            purchaser_member=self.purchaser, approver_member=self.purchaser,
        )
        with self.assertRaises(ValidationError):
            bon.full_clean()

    def test_valid_bon_saves_successfully(self):
        bon = BonDeCommande.objects.create(
            house=self.house, budget_year=self.budget_year, number="BB260001",
            purchase_date=date(2026, 1, 7), short_description="Nettoyants",
            total=Decimal("19.49"), sub_budget=self.sub_budget,
            purchaser_member=self.purchaser, approver_member=self.approver,
        )
        self.assertEqual(bon.status, BonStatus.DRAFT)
        self.assertIsNotNone(bon.pk)

    def test_snapshot_fields_captured(self):
        bon = BonDeCommande(
            house=self.house, budget_year=self.budget_year, number="BB260001",
            purchase_date=date(2026, 1, 7), short_description="Test",
            total=Decimal("19.49"), sub_budget=self.sub_budget,
            purchaser_member=self.purchaser, purchaser_apartment=self.apt,
            approver_member=self.approver,
        )
        bon.refresh_snapshot_fields()
        self.assertEqual(bon.purchaser_name_snapshot, "Marylin Lamarche")
        self.assertEqual(bon.purchaser_unit_snapshot, "202")
        self.assertEqual(bon.approver_name_snapshot, "Carl-David Fortin")

    def test_sub_budget_must_match_budget_year(self):
        other_year = BudgetYear.objects.create(
            house=self.house, year=2027, annual_budget_total=Decimal("10000.00")
        )
        other_sb = SubBudget.objects.create(
            budget_year=other_year, trace_code=7,
            name="Produits ménager", planned_amount=Decimal("300.00")
        )
        bon = BonDeCommande(
            house=self.house, budget_year=self.budget_year, number="BB260001",
            purchase_date=date(2026, 1, 7), short_description="Test",
            total=Decimal("10.00"), sub_budget=other_sb,
            purchaser_member=self.purchaser,
        )
        with self.assertRaises(ValidationError):
            bon.full_clean()


class ReceiptOcrParseTests(TestCase):
    """Test GPT receipt analysis response parsing (batch format)."""

    def test_parse_batch_typical(self):
        raw = '[{"filename": "recu1.png", "merchant": "DOLLARAMA", "purchase_date": "2025-01-15", "subtotal": 3.75, "tps": 0.19, "tvq": 0.37, "total": 4.31, "member_name": "Abla", "apartment_number": "207"}]'
        results = ReceiptOcrService._parse_batch_response(raw, ["recu1.png"])
        self.assertEqual(len(results), 1)
        r = results[0]
        self.assertEqual(r["filename"], "recu1.png")
        self.assertEqual(r["merchant"], "DOLLARAMA")
        self.assertEqual(r["purchase_date"], date(2025, 1, 15))
        self.assertEqual(r["subtotal"], Decimal("3.75"))
        self.assertEqual(r["tps"], Decimal("0.19"))
        self.assertEqual(r["tvq"], Decimal("0.37"))
        self.assertEqual(r["total"], Decimal("4.31"))
        self.assertEqual(r["member_name"], "Abla")
        self.assertEqual(r["apartment_number"], "207")

    def test_parse_batch_normalizes_field_confidence_scores(self):
        raw = json.dumps([{
            "filename": "recu1.png",
            "merchant": "DOLLARAMA",
            "purchase_date": "2025-01-15",
            "total": 4.31,
            "field_confidence_scores": {
                "merchant": 8,
                "purchase_date": "7",
                "total": 9,
            },
        }])
        results = ReceiptOcrService._parse_batch_response(raw, ["recu1.png"])
        scores = results[0]["field_confidence_scores"]
        self.assertEqual(scores["merchant"], 8)
        self.assertEqual(scores["purchase_date"], 7)
        self.assertEqual(scores["total"], 9)
        self.assertEqual(scores["member_name"], "NA")
        self.assertEqual(scores["summary"], "NA")

    def test_parse_batch_empty(self):
        results = ReceiptOcrService._parse_batch_response("", ["a.png"])
        self.assertEqual(len(results), 1)
        self.assertEqual(results[0]["merchant"], "")
        self.assertEqual(results[0]["member_name"], "")
        self.assertIsNone(results[0]["purchase_date"])
        self.assertIsNone(results[0]["total"])

    def test_parse_batch_with_nulls(self):
        raw = '[{"filename": "r.png", "merchant": "METRO", "purchase_date": null, "subtotal": null, "tps": null, "tvq": null, "total": 12.34, "member_name": "ILLISIBLE", "apartment_number": "206"}]'
        results = ReceiptOcrService._parse_batch_response(raw, ["r.png"])
        r = results[0]
        self.assertEqual(r["merchant"], "METRO")
        self.assertIsNone(r["purchase_date"])
        self.assertIsNone(r["subtotal"])
        self.assertEqual(r["total"], Decimal("12.34"))
        self.assertEqual(r["member_name"], "ILLISIBLE")
        self.assertEqual(r["apartment_number"], "206")

    def test_parse_batch_with_code_fences(self):
        raw = '```json\n[{"filename": "r.png", "merchant": "RONA", "purchase_date": "2025-03-01", "subtotal": 45.00, "tps": 2.25, "tvq": 4.49, "total": 51.74, "member_name": "Carl", "apartment_number": "101"}]\n```'
        results = ReceiptOcrService._parse_batch_response(raw, ["r.png"])
        self.assertEqual(results[0]["merchant"], "RONA")
        self.assertEqual(results[0]["tps"], Decimal("2.25"))
        self.assertEqual(results[0]["total"], Decimal("51.74"))

    def test_parse_batch_single_object_fallback(self):
        """GPT might return a single object instead of array for 1 receipt."""
        raw = '{"filename": "x.png", "merchant": "IGA", "total": 9.99, "purchase_date": null, "subtotal": null, "tps": null, "tvq": null, "member_name": "", "apartment_number": ""}'
        results = ReceiptOcrService._parse_batch_response(raw, ["x.png"])
        self.assertEqual(len(results), 1)
        self.assertEqual(results[0]["merchant"], "IGA")
        self.assertEqual(results[0]["total"], Decimal("9.99"))

    def test_parse_batch_multiple_receipts(self):
        raw = json.dumps([
            {"filename": "a.png", "merchant": "Metro", "total": 10.00, "purchase_date": None, "subtotal": None, "tps": None, "tvq": None, "member_name": "A", "apartment_number": "101"},
            {"filename": "b.png", "merchant": "IGA", "total": 20.00, "purchase_date": None, "subtotal": None, "tps": None, "tvq": None, "member_name": "B", "apartment_number": "202"},
        ])
        results = ReceiptOcrService._parse_batch_response(raw, ["a.png", "b.png"])
        self.assertEqual(len(results), 2)
        self.assertEqual(results[0]["merchant"], "Metro")
        self.assertEqual(results[1]["merchant"], "IGA")

    def test_parse_batch_paper_bc_signers_and_ambiguity(self):
        raw = json.dumps([{
            "filename": "BC16011.pdf - Page 1",
            "document_type": "paper_bc",
            "bc_number": "16011",
            "expense_member_name": "Marylin Lamarche",
            "expense_apartment": "202",
            "validator_member_name": "René Côté",
            "validator_apartment": "203",
            "signer_roles_ambiguous": True,
            "total": 547.93,
        }])
        results = ReceiptOcrService._parse_batch_response(raw, ["BC16011.pdf"])
        self.assertEqual(results[0]["document_type"], "paper_bc")
        self.assertEqual(results[0]["expense_member_name"], "Marylin Lamarche")
        self.assertEqual(results[0]["validator_member_name"], "René Côté")
        self.assertEqual(results[0]["validator_apartment"], "203")
        self.assertTrue(results[0]["signer_roles_ambiguous"])

    def test_is_available_without_key(self):
        with self.settings(OPENAI_API_KEY=""):
            self.assertFalse(ReceiptOcrService.is_available())

    def test_is_available_with_key(self):
        with self.settings(OPENAI_API_KEY="sk-test-key"):
            self.assertTrue(ReceiptOcrService.is_available())

    def test_safe_decimal(self):
        self.assertEqual(ReceiptOcrService._safe_decimal(12.50), Decimal("12.50"))
        self.assertEqual(ReceiptOcrService._safe_decimal("3.75"), Decimal("3.75"))
        self.assertIsNone(ReceiptOcrService._safe_decimal(None))
        self.assertIsNone(ReceiptOcrService._safe_decimal("abc"))

    def test_safe_date(self):
        self.assertEqual(ReceiptOcrService._safe_date("2025-01-15"), date(2025, 1, 15))
        self.assertIsNone(ReceiptOcrService._safe_date(""))
        self.assertIsNone(ReceiptOcrService._safe_date("not-a-date"))
        self.assertIsNone(ReceiptOcrService._safe_date(None))

    def test_parse_paper_bc(self):
        """Paper BC detection: GPT returns document_type=paper_bc with bc_number."""
        raw = json.dumps([
            {
                "filename": "BC16011.pdf - Page 1",
                "document_type": "paper_bc",
                "bc_number": "16011",
                "associated_bc_number": "",
                "supplier_name": "Gicleurs de l'Estrie",
                "supplier_address": "1110 Bélanger, Sherbrooke",
                "member_name": "",
                "apartment_number": "",
                "merchant": "",
                "purchase_date": "2024-11-29",
                "subtotal": 476.56,
                "tps": 23.83,
                "tvq": 47.54,
                "total": 547.93,
                "summary": "Plomberie - installation manomètre",
            },
            {
                "filename": "BC16011.pdf - Page 2",
                "document_type": "invoice",
                "bc_number": "",
                "associated_bc_number": "16011",
                "supplier_name": "Gicleurs de l'Estrie inc.",
                "supplier_address": "",
                "member_name": "",
                "apartment_number": "",
                "merchant": "",
                "purchase_date": "2024-11-29",
                "subtotal": 476.56,
                "tps": 23.83,
                "tvq": 47.54,
                "total": 547.93,
                "summary": "Facture matériaux et main d'oeuvre",
            },
        ])
        results = ReceiptOcrService._parse_batch_response(
            raw, ["BC16011.pdf - Page 1", "BC16011.pdf - Page 2"]
        )
        self.assertEqual(len(results), 2)
        # Paper BC
        self.assertEqual(results[0]["document_type"], "paper_bc")
        self.assertEqual(results[0]["bc_number"], "16011")
        self.assertEqual(results[0]["supplier_name"], "Gicleurs de l'Estrie")
        self.assertEqual(results[0]["total"], Decimal("547.93"))
        # Invoice
        self.assertEqual(results[1]["document_type"], "invoice")
        self.assertEqual(results[1]["associated_bc_number"], "16011")
        self.assertEqual(results[1]["supplier_name"], "Gicleurs de l'Estrie inc.")

    def test_parse_mixed_upload(self):
        """Mixed upload: receipt + paper BC in same batch."""
        raw = json.dumps([
            {
                "filename": "receipt.png",
                "document_type": "receipt",
                "bc_number": "",
                "associated_bc_number": "",
                "supplier_name": "",
                "supplier_address": "",
                "member_name": "Jean Dupont",
                "apartment_number": "305",
                "merchant": "RONA",
                "purchase_date": "2025-03-01",
                "subtotal": 45.00,
                "tps": 2.25,
                "tvq": 4.49,
                "total": 51.74,
                "summary": "Vis et peinture",
            },
            {
                "filename": "BC16739.pdf - Page 1",
                "document_type": "paper_bc",
                "bc_number": "16739",
                "associated_bc_number": "",
                "supplier_name": "Produits Sany",
                "supplier_address": "",
                "member_name": "",
                "apartment_number": "",
                "merchant": "",
                "purchase_date": "2026-01-07",
                "subtotal": 19.49,
                "tps": None,
                "tvq": None,
                "total": 19.49,
                "summary": "NUBIOCAL 900ML",
            },
        ])
        results = ReceiptOcrService._parse_batch_response(
            raw, ["receipt.png", "BC16739.pdf - Page 1"]
        )
        self.assertEqual(len(results), 2)
        # Receipt
        self.assertEqual(results[0]["document_type"], "receipt")
        self.assertEqual(results[0]["member_name"], "Jean Dupont")
        self.assertEqual(results[0]["apartment_number"], "305")
        self.assertEqual(results[0]["merchant"], "RONA")
        # Paper BC
        self.assertEqual(results[1]["document_type"], "paper_bc")
        self.assertEqual(results[1]["bc_number"], "16739")
        self.assertEqual(results[1]["supplier_name"], "Produits Sany")

    def test_parse_defaults_to_receipt(self):
        """Old-format response without document_type defaults to receipt."""
        raw = '[{"filename": "old.png", "merchant": "IGA", "total": 5.00}]'
        results = ReceiptOcrService._parse_batch_response(raw, ["old.png"])
        self.assertEqual(results[0]["document_type"], "receipt")
        self.assertEqual(results[0]["bc_number"], "")
        self.assertEqual(results[0]["associated_bc_number"], "")


class OcrMemberDirectoryTests(TestCase):
    def setUp(self):
        self.house = House.objects.create(
            code="BB", name="Maison BB", account_number="13-51200"
        )
        self.budget_year = BudgetYear.objects.create(
            house=self.house, year=2026, annual_budget_total=Decimal("1000.00")
        )
        self.sub_budget = SubBudget.objects.create(
            budget_year=self.budget_year,
            trace_code=7,
            name="Produits ménager",
            planned_amount=Decimal("100.00"),
        )
        self.member = Member.objects.create(first_name="Serge", last_name="Laroche")
        self.apartment = Apartment.objects.create(house=self.house, code="105")
        Residency.objects.create(
            member=self.member,
            apartment=self.apartment,
            start_date=date(2020, 1, 1),
        )

    def test_batch_prompt_includes_canonical_member_directory(self):
        prompt = ReceiptOcrService._build_batch_prompt(self.house)
        self.assertIn("RÉPERTOIRE OFFICIEL DES MEMBRES ACTIFS", prompt)
        self.assertIn("Appartement 105: Serge Laroche", prompt)
        self.assertIn("retourne EXACTEMENT le nom officiel", prompt)
        self.assertIn("le 2e signataire peut etre une personne EXTERNE", prompt)

    def test_resolve_member_assignment_prefers_fuzzy_name_match(self):
        other_member = Member.objects.create(first_name="Pierre", last_name="Bouchard")
        other_apartment = Apartment.objects.create(house=self.house, code="205")
        Residency.objects.create(
            member=other_member,
            apartment=other_apartment,
            start_date=date(2020, 1, 1),
        )

        apartment, member = _resolve_member_assignment(self.house, "205", "Serge Laroch")

        self.assertEqual(member, self.member)
        self.assertEqual(apartment, self.apartment)

    def test_paper_bc_signer_initials_prefill_fuzzy_member_and_canonical_apartment(self):
        fake_receipt = ReceiptFile.objects.create(
            bon_de_commande=BonDeCommande.objects.create(
                house=self.house,
                budget_year=self.budget_year,
                number="BB260001",
                purchase_date=date(2026, 1, 7),
                short_description="Test",
                total=Decimal("10.00"),
                sub_budget=self.sub_budget,
                purchaser_member=self.member,
            ),
            file=SimpleUploadedFile("BC17186.pdf", b"fake pdf", content_type="application/pdf"),
            original_filename="BC17186.pdf",
            content_type="application/pdf",
        )
        ef = ReceiptExtractedFields.objects.create(
            receipt_file=fake_receipt,
            document_type_candidate="paper_bc",
            expense_member_name_candidate="Serge Laroch",
            expense_apartment_candidate="",
        )

        initial, purchaser_mismatch, _, purchaser_apartment, _ = _paper_bc_signer_initials(
            self.house,
            ef,
        )

        self.assertEqual(initial["expense_member"], self.member.pk)
        self.assertEqual(initial["expense_apartment"], "105")
        self.assertFalse(purchaser_mismatch)
        self.assertEqual(purchaser_apartment, self.apartment)


class OcrBatchSplittingTests(TestCase):
    """Test that PDFs are isolated into their own batches."""

    def test_each_pdf_gets_own_batch(self):
        file_map = {
            "BC16011.pdf": "/tmp/BC16011.pdf",
            "BC16739.pdf": "/tmp/BC16739.pdf",
        }
        # Mock page counts so each PDF has 2 pages
        with patch.object(ReceiptOcrService, "_count_pages", return_value={
            "BC16011.pdf": 2, "BC16739.pdf": 2,
        }):
            batches = ReceiptOcrService._split_file_map(file_map)
        self.assertEqual(len(batches), 2)
        self.assertIn("BC16011.pdf", batches[0])
        self.assertIn("BC16739.pdf", batches[1])

    def test_images_grouped_together(self):
        file_map = {
            "receipt1.png": "/tmp/receipt1.png",
            "receipt2.jpg": "/tmp/receipt2.jpg",
            "receipt3.png": "/tmp/receipt3.png",
        }
        with patch.object(ReceiptOcrService, "_count_pages", return_value={
            "receipt1.png": 1, "receipt2.jpg": 1, "receipt3.png": 1,
        }):
            batches = ReceiptOcrService._split_file_map(file_map)
        self.assertEqual(len(batches), 1)
        self.assertEqual(len(batches[0]), 3)

    def test_pdf_separates_from_images(self):
        file_map = {
            "receipt1.png": "/tmp/receipt1.png",
            "BC16011.pdf": "/tmp/BC16011.pdf",
            "receipt2.png": "/tmp/receipt2.png",
        }
        with patch.object(ReceiptOcrService, "_count_pages", return_value={
            "receipt1.png": 1, "BC16011.pdf": 3, "receipt2.png": 1,
        }):
            batches = ReceiptOcrService._split_file_map(file_map)
        self.assertEqual(len(batches), 3)
        self.assertIn("receipt1.png", batches[0])
        self.assertIn("BC16011.pdf", batches[1])
        self.assertIn("receipt2.png", batches[2])

    def test_images_respect_page_limit(self):
        file_map = {f"r{i}.png": f"/tmp/r{i}.png" for i in range(6)}
        with patch.object(ReceiptOcrService, "_count_pages", return_value={
            f"r{i}.png": 1 for i in range(6)
        }):
            batches = ReceiptOcrService._split_file_map(file_map)
        # 6 images at 1 page each, MAX=4 → 2 batches (4+2)
        self.assertEqual(len(batches), 2)
        self.assertEqual(len(batches[0]), 4)
        self.assertEqual(len(batches[1]), 2)


class PaperBcFinalizationTests(TestCase):
    """Test that _finalize_bons correctly handles paper BC + invoice + receipt mixes."""

    def setUp(self):
        from bons.models import ReceiptFile, ReceiptExtractedFields
        self.house = House.objects.create(code="BB", name="Maison BB", account_number="13-51200")
        self.budget_year = BudgetYear.objects.create(
            house=self.house, year=2026, annual_budget_total=Decimal("12237.00")
        )
        self.sub_budget = SubBudget.objects.create(
            budget_year=self.budget_year, trace_code=7,
            name="Produits ménager", planned_amount=Decimal("300.00")
        )
        self.member = Member.objects.create(first_name="Marylin", last_name="Lamarche")
        self.apt = Apartment.objects.create(house=self.house, code="202")
        Residency.objects.create(
            member=self.member, apartment=self.apt, start_date=date(2020, 1, 1)
        )
        self.validator_member = Member.objects.create(first_name="René", last_name="Côté")
        self.validator_apt = Apartment.objects.create(house=self.house, code="203")
        Residency.objects.create(
            member=self.validator_member, apartment=self.validator_apt, start_date=date(2020, 1, 1)
        )
        self.treasurer_member = Member.objects.create(first_name="Trésorier", last_name="Test")
        self.treasurer_apt = Apartment.objects.create(house=self.house, code="204")
        Residency.objects.create(
            member=self.treasurer_member, apartment=self.treasurer_apt, start_date=date(2020, 1, 1)
        )
        self.user = User.objects.create_user(
            username="tresorier", password="test123", role=20,
            member=self.treasurer_member,
            house=self.house,
        )

        # Create a scan session bon
        self.scan_session = BonDeCommande()
        self.scan_session.house = self.house
        self.scan_session.budget_year = self.budget_year
        self.scan_session.number = "BB260099"
        self.scan_session.purchase_date = date(2026, 1, 7)
        self.scan_session.short_description = "(scan session)"
        self.scan_session.total = 0
        self.scan_session.sub_budget = self.sub_budget
        self.scan_session.purchaser_member = self.treasurer_member
        self.scan_session.created_by = self.user
        self.scan_session.status = BonStatus.READY_FOR_REVIEW
        self.scan_session.is_scan_session = True
        super(BonDeCommande, self.scan_session).save()

        # Create receipt files attached to scan session
        self.receipt_bc = ReceiptFile.objects.create(
            bon_de_commande=self.scan_session,
            original_filename="BC16011.pdf",
            content_type="application/pdf",
        )
        self.receipt_inv = ReceiptFile.objects.create(
            bon_de_commande=self.scan_session,
            original_filename="facture.png",
            content_type="image/png",
        )
        self.receipt_regular = ReceiptFile.objects.create(
            bon_de_commande=self.scan_session,
            original_filename="recu_rona.jpg",
            content_type="image/jpeg",
        )

        # Extracted fields for paper BC
        ReceiptExtractedFields.objects.create(
            receipt_file=self.receipt_bc,
            document_type_candidate="paper_bc",
            final_document_type="paper_bc",
            bc_number_candidate="16011",
            final_bc_number="16011",
            supplier_name_candidate="Gicleurs de l'Estrie",
            final_supplier_name="Gicleurs de l'Estrie",
            total_candidate=Decimal("547.93"),
            final_total=Decimal("547.93"),
            purchase_date_candidate=date(2024, 11, 29),
            final_purchase_date=date(2024, 11, 29),
            expense_member_name_candidate="Maryline Lamarche",
            final_expense_member_name="Maryline Lamarche",
            expense_apartment_candidate="202",
            final_expense_apartment="202",
            validator_member_name_candidate="René Côté",
            final_validator_member_name="René Côté",
            validator_apartment_candidate="203",
            final_validator_apartment="203",
            signer_roles_ambiguous_candidate=True,
            signer_roles_ambiguous_final=True,
            sub_budget=self.sub_budget,
            summary_candidate="Installation purgeur",
            final_summary="Installation purgeur",
        )
        # Extracted fields for invoice linked to paper BC
        ReceiptExtractedFields.objects.create(
            receipt_file=self.receipt_inv,
            document_type_candidate="invoice",
            final_document_type="invoice",
            associated_bc_number_candidate="16011",
            final_associated_bc_number="16011",
            supplier_name_candidate="Gicleurs de l'Estrie inc.",
            final_supplier_name="Gicleurs de l'Estrie inc.",
            total_candidate=Decimal("547.93"),
            final_total=Decimal("547.93"),
            summary_candidate="Travaux plomberie",
            final_summary="Travaux plomberie",
            sub_budget=self.sub_budget,
        )
        # Extracted fields for regular receipt
        ReceiptExtractedFields.objects.create(
            receipt_file=self.receipt_regular,
            document_type_candidate="receipt",
            final_document_type="receipt",
            member_name_candidate="Marylin Lamarche",
            final_member_name="Marylin Lamarche",
            apartment_number_candidate="202",
            final_apartment_number="202",
            merchant_candidate="RONA",
            final_merchant="RONA",
            total_candidate=Decimal("25.00"),
            final_total=Decimal("25.00"),
            summary_candidate="Vis et peinture",
            final_summary="Vis et peinture",
            sub_budget=self.sub_budget,
        )

    def test_finalize_creates_paper_bc_and_regular_bons(self):
        """Mixed upload: paper BC with invoice + regular receipt → 2 bons."""
        from django.test import RequestFactory
        from bons.views import OcrReviewView

        factory = RequestFactory()
        request = factory.get("/")
        request.user = self.user
        # Attach message storage
        from django.contrib.messages.storage.fallback import FallbackStorage
        setattr(request, "session", "session")
        setattr(request, "_messages", FallbackStorage(request))

        view = OcrReviewView()
        response = view._finalize_bons(request, self.scan_session)

        # Should have created 2 bons: 1 paper BC + 1 regular
        created_bons = BonDeCommande.objects.filter(
            is_scan_session=False
        ).exclude(status=BonStatus.VOID)
        self.assertEqual(created_bons.count(), 2)

        # Paper BC bon
        paper_bon = created_bons.filter(is_paper_bc=True).first()
        self.assertIsNotNone(paper_bon)
        self.assertEqual(paper_bon.number, "16011")
        self.assertEqual(paper_bon.paper_bc_number, "16011")
        self.assertEqual(paper_bon.status, BonStatus.READY_FOR_VALIDATION)
        self.assertEqual(paper_bon.merchant_name, "Gicleurs de l'Estrie")
        self.assertNotEqual(paper_bon.purchaser_member, self.treasurer_member)
        self.assertEqual(paper_bon.purchaser_member, self.member)
        self.assertEqual(paper_bon.purchaser_apartment, self.apt)
        self.assertEqual(paper_bon.approver_member, self.validator_member)
        self.assertEqual(paper_bon.approver_apartment, self.validator_apt)
        # Paper BC bon should have 2 receipt files (BC + invoice)
        self.assertEqual(paper_bon.receipt_files.count(), 2)

        # Regular bon
        regular_bon = created_bons.filter(is_paper_bc=False).first()
        self.assertIsNotNone(regular_bon)
        self.assertTrue(regular_bon.number.startswith("BB26"))
        self.assertEqual(regular_bon.receipt_files.count(), 1)

        # Scan session should be voided
        self.scan_session.refresh_from_db()
        self.assertEqual(self.scan_session.status, BonStatus.VOID)

    def test_finalize_handles_duplicate_paper_bc_number(self):
        """If a paper BC number already exists, append suffix."""
        # Create an existing bon with number 16011
        BonDeCommande.objects.create(
            house=self.house, budget_year=self.budget_year,
            number="16011", purchase_date=date(2026, 1, 1),
            short_description="Existing", total=Decimal("100"),
            sub_budget=self.sub_budget, purchaser_member=self.member,
        )

        from django.test import RequestFactory
        from bons.views import OcrReviewView
        from django.contrib.messages.storage.fallback import FallbackStorage

        factory = RequestFactory()
        request = factory.get("/")
        request.user = self.user
        setattr(request, "session", "session")
        setattr(request, "_messages", FallbackStorage(request))

        view = OcrReviewView()
        view._finalize_bons(request, self.scan_session)

        # Should use suffix since 16011 is taken
        paper_bon = BonDeCommande.objects.filter(
            is_paper_bc=True, is_scan_session=False
        ).exclude(status=BonStatus.VOID).first()
        self.assertIsNotNone(paper_bon)
        self.assertEqual(paper_bon.number, "16011-2")
        self.assertEqual(paper_bon.paper_bc_number, "16011")

    def test_finalize_reactivates_voided_paper_bc_instead_of_creating_new_one(self):
        from django.contrib.messages.storage.fallback import FallbackStorage
        from django.test import RequestFactory
        from bons.views import OcrReviewView

        archived_bon = BonDeCommande.objects.create(
            house=self.house,
            budget_year=self.budget_year,
            number="16011",
            purchase_date=date(2026, 1, 1),
            short_description="Ancien BC annulé",
            total=Decimal("100.00"),
            sub_budget=self.sub_budget,
            purchaser_member=self.member,
            status=BonStatus.VOID,
            is_paper_bc=True,
            paper_bc_number="16011",
            void_reason="Test",
            voided_at=timezone.now(),
        )
        old_receipt = ReceiptFile.objects.create(
            bon_de_commande=archived_bon,
            file=SimpleUploadedFile("old.png", b"old image", content_type="image/png"),
            original_filename="old.png",
            content_type="image/png",
            ocr_status=OcrStatus.CORRECTED,
        )

        request = RequestFactory().get("/")
        request.user = self.user
        setattr(request, "session", "session")
        setattr(request, "_messages", FallbackStorage(request))

        OcrReviewView()._finalize_bons(request, self.scan_session)

        archived_bon.refresh_from_db()
        old_receipt.refresh_from_db()
        self.assertEqual(archived_bon.status, BonStatus.READY_FOR_VALIDATION)
        self.assertIsNone(archived_bon.voided_at)
        self.assertEqual(archived_bon.paper_bc_number, "16011")
        self.assertTrue(old_receipt.is_archived)
        self.assertEqual(archived_bon.active_receipt_files.count(), 2)
        self.assertSetEqual(
            set(archived_bon.active_receipt_files.values_list("original_filename", flat=True)),
            {"BC16011.pdf", "facture.png"},
        )
        self.assertEqual(
            BonDeCommande.objects.filter(
                is_scan_session=False,
                is_paper_bc=True,
                paper_bc_number="16011",
            ).count(),
            1,
        )

    def test_orphan_invoices_grouped_as_regular(self):
        """Invoices with no matching paper BC are treated as regular receipts."""
        from bons.models import ReceiptExtractedFields
        # Change the invoice to point to a non-existent BC
        ef = self.receipt_inv.extracted_fields
        ef.final_associated_bc_number = "99999"
        ef.associated_bc_number_candidate = "99999"
        ef.save()

        from django.test import RequestFactory
        from bons.views import OcrReviewView
        from django.contrib.messages.storage.fallback import FallbackStorage

        factory = RequestFactory()
        request = factory.get("/")
        request.user = self.user
        setattr(request, "session", "session")
        setattr(request, "_messages", FallbackStorage(request))

        view = OcrReviewView()
        view._finalize_bons(request, self.scan_session)

        # Paper BC bon (BC alone, no matched invoices)
        paper_bon = BonDeCommande.objects.filter(
            is_paper_bc=True, is_scan_session=False
        ).exclude(status=BonStatus.VOID).first()
        self.assertIsNotNone(paper_bon)
        self.assertEqual(paper_bon.receipt_files.count(), 1)  # Just the BC

        # Regular bons should have the orphan invoice + regular receipt
        regular_bons = BonDeCommande.objects.filter(
            is_paper_bc=False, is_scan_session=False
        ).exclude(status=BonStatus.VOID)
        total_receipts = sum(b.receipt_files.count() for b in regular_bons)
        self.assertEqual(total_receipts, 2)  # invoice + receipt

    def test_finalize_leaves_approver_empty_when_no_second_signer(self):
        ef = self.receipt_bc.extracted_fields
        ef.final_validator_member_name = ""
        ef.final_validator_apartment = ""
        ef.validator_member_name_candidate = ""
        ef.validator_apartment_candidate = ""
        ef.save()

        from django.test import RequestFactory
        from bons.views import OcrReviewView
        from django.contrib.messages.storage.fallback import FallbackStorage

        request = RequestFactory().get("/")
        request.user = self.user
        setattr(request, "session", "session")
        setattr(request, "_messages", FallbackStorage(request))

        OcrReviewView()._finalize_bons(request, self.scan_session)
        paper_bon = BonDeCommande.objects.filter(
            is_paper_bc=True, is_scan_session=False,
        ).exclude(status=BonStatus.VOID).first()
        self.assertIsNotNone(paper_bon)
        self.assertIsNone(paper_bon.approver_member)
        self.assertIsNone(paper_bon.approver_apartment)

    def test_finalize_prefers_invoice_amounts_and_derives_missing_taxes(self):
        self.receipt_bc.extracted_fields.final_subtotal = Decimal("100.00")
        self.receipt_bc.extracted_fields.final_tps = None
        self.receipt_bc.extracted_fields.final_tvq = None
        self.receipt_bc.extracted_fields.final_total = Decimal("100.00")
        self.receipt_bc.extracted_fields.save()

        self.receipt_inv.extracted_fields.final_subtotal = Decimal("100.00")
        self.receipt_inv.extracted_fields.final_tps = None
        self.receipt_inv.extracted_fields.final_tvq = None
        self.receipt_inv.extracted_fields.final_total = Decimal("114.98")
        self.receipt_inv.extracted_fields.save()

        from django.test import RequestFactory
        from bons.views import OcrReviewView
        from django.contrib.messages.storage.fallback import FallbackStorage

        request = RequestFactory().get("/")
        request.user = self.user
        setattr(request, "session", "session")
        setattr(request, "_messages", FallbackStorage(request))

        OcrReviewView()._finalize_bons(request, self.scan_session)
        paper_bon = BonDeCommande.objects.filter(
            is_paper_bc=True,
            is_scan_session=False,
        ).exclude(status=BonStatus.VOID).first()

        self.assertIsNotNone(paper_bon)
        self.assertEqual(paper_bon.subtotal, Decimal("100.00"))
        self.assertEqual(paper_bon.tps, Decimal("5.00"))
        self.assertEqual(paper_bon.tvq, Decimal("9.98"))

    def test_finalize_preserves_untaxed_extra_amount(self):
        self.receipt_bc.extracted_fields.final_subtotal = Decimal("100.00")
        self.receipt_bc.extracted_fields.final_tps = None
        self.receipt_bc.extracted_fields.final_tvq = None
        self.receipt_bc.extracted_fields.final_total = Decimal("124.98")
        self.receipt_bc.extracted_fields.save()

        self.receipt_inv.extracted_fields.final_subtotal = Decimal("100.00")
        self.receipt_inv.extracted_fields.final_tps = None
        self.receipt_inv.extracted_fields.final_tvq = None
        self.receipt_inv.extracted_fields.final_untaxed_extra_amount = Decimal("10.00")
        self.receipt_inv.extracted_fields.final_total = Decimal("124.98")
        self.receipt_inv.extracted_fields.save()

        from django.test import RequestFactory
        from bons.views import OcrReviewView
        from django.contrib.messages.storage.fallback import FallbackStorage

        request = RequestFactory().get("/")
        request.user = self.user
        setattr(request, "session", "session")
        setattr(request, "_messages", FallbackStorage(request))

        OcrReviewView()._finalize_bons(request, self.scan_session)
        paper_bon = BonDeCommande.objects.filter(
            is_paper_bc=True,
            is_scan_session=False,
        ).exclude(status=BonStatus.VOID).first()

        self.assertIsNotNone(paper_bon)
        self.assertEqual(paper_bon.subtotal, Decimal("100.00"))
        self.assertEqual(paper_bon.tps, Decimal("5.00"))
        self.assertEqual(paper_bon.tvq, Decimal("9.98"))
        self.assertEqual(paper_bon.untaxed_extra_amount, Decimal("10.00"))
        self.assertEqual(paper_bon.total, Decimal("124.98"))


class MobileCaptureFlowTests(TestCase):
    def setUp(self):
        self.house = House.objects.create(code="BB", name="Maison BB", account_number="13-51200")
        self.budget_year = BudgetYear.objects.create(
            house=self.house,
            year=2026,
            annual_budget_total=Decimal("12237.00"),
        )
        self.sub_budget = SubBudget.objects.create(
            budget_year=self.budget_year,
            trace_code=7,
            name="Produits ménager",
            planned_amount=Decimal("300.00"),
        )
        self.member = Member.objects.create(first_name="Marylin", last_name="Lamarche")
        self.apartment = Apartment.objects.create(house=self.house, code="202")
        Residency.objects.create(
            member=self.member,
            apartment=self.apartment,
            start_date=date(2020, 1, 1),
        )
        self.user = User.objects.create_user(
            username="mobile-tresorier",
            password="test123",
            role=User.Role.TREASURER,
            house=self.house,
            member=self.member,
        )
        self.client.login(username="mobile-tresorier", password="test123")

    @staticmethod
    def _mobile_photo(name="capture.jpg"):
        return SimpleUploadedFile(name, b"fake image data", content_type="image/jpeg")

    def _set_mobile_capture_session(self, scan_session):
        session = self.client.session
        session["mobile_capture_scan_session_id"] = scan_session.pk
        session.save()

    @patch("bons.ocr_service.ReceiptOcrService.is_available", return_value=True)
    def test_mobile_capture_adds_photo_to_scan_session(self, mock_ocr_available):
        response = self.client.post(
            reverse("bons:mobile-capture"),
            {
                "budget_year": self.budget_year.pk,
                "photo": self._mobile_photo(),
            },
        )

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response.headers["Location"], reverse("bons:mobile-capture"))
        scan_session = BonDeCommande.objects.get(is_scan_session=True)
        self.assertEqual(scan_session.status, BonStatus.DRAFT)
        self.assertEqual(scan_session.active_receipt_files.count(), 1)
        self.assertEqual(
            self.client.session["mobile_capture_scan_session_id"],
            scan_session.pk,
        )

    @patch("bons.ocr_service.ReceiptOcrService.process_receipts_batch")
    @patch("bons.ocr_service.ReceiptOcrService.is_available", return_value=True)
    def test_mobile_capture_finalize_refuses_missing_signature_and_apartment(
        self,
        mock_ocr_available,
        mock_process,
    ):
        scan_session = create_scan_session(user=self.user, budget_year=self.budget_year)
        receipt = ReceiptFile.objects.create(
            bon_de_commande=scan_session,
            file=self._mobile_photo("missing-signature.jpg"),
            original_filename="missing-signature.jpg",
            content_type="image/jpeg",
            uploaded_by=self.user,
        )
        self._set_mobile_capture_session(scan_session)

        def fake_process(receipts, house=None):
            ReceiptExtractedFields.objects.update_or_create(
                receipt_file=receipt,
                defaults={
                    "document_type_candidate": "receipt",
                    "member_name_candidate": "",
                    "apartment_number_candidate": "",
                },
            )
            return receipts, ""

        mock_process.side_effect = fake_process

        response = self.client.post(
            reverse("bons:mobile-capture-finalize"),
            follow=True,
        )

        scan_session.refresh_from_db()
        self.assertEqual(scan_session.status, BonStatus.DRAFT)
        self.assertContains(response, "Traitement refusé")
        self.assertContains(response, "signature lisible")

    @patch("bons.ocr_service.ReceiptOcrService.process_receipts_batch")
    @patch("bons.ocr_service.ReceiptOcrService.is_available", return_value=True)
    def test_mobile_capture_finalize_redirects_to_review_when_signature_found(
        self,
        mock_ocr_available,
        mock_process,
    ):
        scan_session = create_scan_session(user=self.user, budget_year=self.budget_year)
        receipt = ReceiptFile.objects.create(
            bon_de_commande=scan_session,
            file=self._mobile_photo("signed-receipt.jpg"),
            original_filename="signed-receipt.jpg",
            content_type="image/jpeg",
            uploaded_by=self.user,
        )
        self._set_mobile_capture_session(scan_session)

        def fake_process(receipts, house=None):
            ReceiptExtractedFields.objects.update_or_create(
                receipt_file=receipt,
                defaults={
                    "document_type_candidate": "receipt",
                    "member_name_candidate": "Marylin Lamarche",
                    "apartment_number_candidate": "202",
                },
            )
            return receipts, ""

        mock_process.side_effect = fake_process

        response = self.client.post(reverse("bons:mobile-capture-finalize"))

        scan_session.refresh_from_db()
        self.assertEqual(scan_session.status, BonStatus.READY_FOR_REVIEW)
        self.assertEqual(
            response.headers["Location"],
            reverse("bons:review", kwargs={"pk": scan_session.pk}) + "?idx=0",
        )
        self.assertNotIn("mobile_capture_scan_session_id", self.client.session)

    @patch("bons.ocr_service.ReceiptOcrService.process_receipts_batch")
    @patch("bons.ocr_service.ReceiptOcrService.is_available", return_value=True)
    def test_mobile_capture_finalize_allows_invoice_with_signed_paper_bc(
        self,
        mock_ocr_available,
        mock_process,
    ):
        scan_session = create_scan_session(user=self.user, budget_year=self.budget_year)
        paper_bc = ReceiptFile.objects.create(
            bon_de_commande=scan_session,
            file=self._mobile_photo("paper-bc.jpg"),
            original_filename="paper-bc.jpg",
            content_type="image/jpeg",
            uploaded_by=self.user,
        )
        invoice = ReceiptFile.objects.create(
            bon_de_commande=scan_session,
            file=self._mobile_photo("invoice.jpg"),
            original_filename="invoice.jpg",
            content_type="image/jpeg",
            uploaded_by=self.user,
        )
        self._set_mobile_capture_session(scan_session)

        def fake_process(receipts, house=None):
            ReceiptExtractedFields.objects.update_or_create(
                receipt_file=paper_bc,
                defaults={
                    "document_type_candidate": "paper_bc",
                    "bc_number_candidate": "16011",
                    "expense_member_name_candidate": "Marylin Lamarche",
                    "expense_apartment_candidate": "202",
                },
            )
            ReceiptExtractedFields.objects.update_or_create(
                receipt_file=invoice,
                defaults={
                    "document_type_candidate": "invoice",
                    "associated_bc_number_candidate": "16011",
                },
            )
            return receipts, ""

        mock_process.side_effect = fake_process

        response = self.client.post(reverse("bons:mobile-capture-finalize"))

        scan_session.refresh_from_db()
        self.assertEqual(scan_session.status, BonStatus.READY_FOR_REVIEW)
        self.assertEqual(
            response.headers["Location"],
            reverse("bons:review", kwargs={"pk": scan_session.pk}) + "?idx=0",
        )


class MismatchWarningTests(TestCase):
    """Test the _get_mismatch_warning helper."""

    def test_matching_totals_no_warning(self):
        from bons.views import _get_mismatch_warning
        from unittest.mock import MagicMock

        receipt = MagicMock()
        receipt.ocr_raw_text = json.dumps([
            {"document_type": "paper_bc", "bc_number": "16011", "total": 547.93},
            {"document_type": "invoice", "associated_bc_number": "16011", "total": 547.93},
        ])
        self.assertIsNone(_get_mismatch_warning(receipt))

    def test_mismatched_totals_returns_warning(self):
        from bons.views import _get_mismatch_warning
        from unittest.mock import MagicMock

        receipt = MagicMock()
        receipt.ocr_raw_text = json.dumps([
            {"document_type": "paper_bc", "bc_number": "16739", "total": 19.49},
            {"document_type": "invoice", "associated_bc_number": "16739", "total": 16.95},
        ])
        warning = _get_mismatch_warning(receipt)
        self.assertIsNotNone(warning)
        self.assertEqual(warning["bc_number"], "16739")
        self.assertAlmostEqual(float(warning["bc_total"]), 19.49)
        self.assertAlmostEqual(float(warning["invoice_total"]), 16.95)

    def test_mismatch_warning_ignores_invoices_for_other_bc_numbers(self):
        from bons.views import _get_mismatch_warning
        from unittest.mock import MagicMock

        receipt = MagicMock()
        receipt.ocr_raw_text = json.dumps([
            {"document_type": "paper_bc", "bc_number": "16739", "total": 19.49},
            {"document_type": "invoice", "associated_bc_number": "16739", "total": 19.49},
            {"document_type": "paper_bc", "bc_number": "17186", "total": 54.58},
            {"document_type": "invoice", "associated_bc_number": "17186", "total": 54.58},
        ])
        self.assertIsNone(_get_mismatch_warning(receipt))

    def test_no_paper_bc_no_warning(self):
        from bons.views import _get_mismatch_warning
        from unittest.mock import MagicMock

        receipt = MagicMock()
        receipt.ocr_raw_text = json.dumps([
            {"document_type": "receipt", "total": 25.00},
        ])
        self.assertIsNone(_get_mismatch_warning(receipt))

    def test_missing_taxes_on_paper_bc_does_not_warn_when_invoice_has_them(self):
        from bons.views import _get_mismatch_warning
        from unittest.mock import MagicMock

        receipt = MagicMock()
        receipt.ocr_raw_text = json.dumps([
            {
                "document_type": "paper_bc",
                "bc_number": "17186",
                "subtotal": 100.00,
                "tps": None,
                "tvq": None,
                "total": 100.00,
            },
            {
                "document_type": "invoice",
                "associated_bc_number": "17186",
                "subtotal": 100.00,
                "tps": 5.00,
                "tvq": 9.98,
                "total": 114.98,
            },
        ])
        self.assertIsNone(_get_mismatch_warning(receipt))

    def test_tax_only_difference_does_not_warn(self):
        from bons.views import _get_mismatch_warning
        from unittest.mock import MagicMock

        receipt = MagicMock()
        receipt.ocr_raw_text = json.dumps([
            {"document_type": "paper_bc", "bc_number": "16739", "total": 19.49},
            {
                "document_type": "invoice",
                "associated_bc_number": "16739",
                "subtotal": 16.95,
                "tps": None,
                "tvq": None,
                "total": None,
            },
        ])
        self.assertIsNone(_get_mismatch_warning(receipt))

    def test_normalize_derives_taxes_from_total_only(self):
        """When only total is known, derive subtotal and taxes from Quebec rates."""
        result = _normalize_document_amounts(total=Decimal("54.58"))
        self.assertIsNotNone(result["subtotal"])
        self.assertIsNotNone(result["tps"])
        self.assertIsNotNone(result["tvq"])
        self.assertEqual(result["total"], Decimal("54.58"))
        # subtotal + tps + tvq should equal total
        self.assertEqual(
            result["subtotal"] + result["tps"] + result["tvq"],
            result["total"],
        )
        # TPS should be 5% of subtotal
        expected_tps = (result["subtotal"] * Decimal("0.05")).quantize(Decimal("0.01"))
        self.assertEqual(result["tps"], expected_tps)

    def test_normalize_derives_taxable_amounts_excluding_untaxed_extra(self):
        result = _normalize_document_amounts(
            subtotal=Decimal("100.00"),
            tps=None,
            tvq=None,
            total=Decimal("124.98"),
            untaxed_extra_amount=Decimal("10.00"),
        )
        self.assertEqual(result["subtotal"], Decimal("100.00"))
        self.assertEqual(result["tps"], Decimal("5.00"))
        self.assertEqual(result["tvq"], Decimal("9.98"))
        self.assertEqual(result["untaxed_extra_amount"], Decimal("10.00"))
        self.assertEqual(result["total"], Decimal("124.98"))

    def test_unverifiable_when_invoice_amounts_null(self):
        """When invoices exist but all amounts are null, return unverifiable warning."""
        from bons.views import _get_mismatch_warning
        from unittest.mock import MagicMock

        receipt = MagicMock()
        receipt.ocr_raw_text = json.dumps([
            {"document_type": "paper_bc", "bc_number": "16739", "total": 19.49},
            {
                "document_type": "invoice",
                "associated_bc_number": "16739",
                "subtotal": None,
                "tps": None,
                "tvq": None,
                "total": None,
            },
        ])
        warning = _get_mismatch_warning(receipt)
        self.assertIsNotNone(warning)
        self.assertTrue(warning.get("unverifiable"))
        self.assertEqual(warning["bc_number"], "16739")
        self.assertAlmostEqual(float(warning["bc_total"]), 19.49)

    def test_supplement_amounts_from_invoices(self):
        """When paper BC has no taxes, they should be filled from invoice data."""
        from bons.views import _supplement_amounts_from_invoices
        from unittest.mock import MagicMock

        receipt = MagicMock()
        receipt.ocr_raw_text = json.dumps([
            {
                "document_type": "paper_bc",
                "bc_number": "17186",
                "subtotal": None,
                "tps": None,
                "tvq": None,
                "total": 54.58,
            },
            {
                "document_type": "invoice",
                "associated_bc_number": "17186",
                "subtotal": 47.47,
                "tps": 2.37,
                "tvq": 4.74,
                "untaxed_extra_amount": 10.00,
                "total": 54.58,
            },
        ])
        initial = {
            "subtotal": None,
            "tps": None,
            "tvq": None,
            "untaxed_extra_amount": None,
            "total": Decimal("54.58"),
        }
        _supplement_amounts_from_invoices(initial, receipt)
        self.assertAlmostEqual(float(initial["subtotal"]), 47.47)
        self.assertAlmostEqual(float(initial["tps"]), 2.37)
        self.assertAlmostEqual(float(initial["tvq"]), 4.74)
        self.assertAlmostEqual(float(initial["untaxed_extra_amount"]), 10.00)
        # total should remain the BC value since it was already set
        self.assertAlmostEqual(float(initial["total"]), 54.58)

    def test_supplement_ignores_invoice_amounts_for_other_bc_numbers(self):
        from bons.views import _supplement_amounts_from_invoices
        from unittest.mock import MagicMock

        receipt = MagicMock()
        receipt.ocr_raw_text = json.dumps([
            {
                "document_type": "paper_bc",
                "bc_number": "17186",
                "subtotal": None,
                "tps": None,
                "tvq": None,
                "total": 54.58,
            },
            {
                "document_type": "invoice",
                "associated_bc_number": "17186",
                "subtotal": 47.47,
                "tps": 2.37,
                "tvq": 4.74,
                "total": 54.58,
            },
            {
                "document_type": "paper_bc",
                "bc_number": "16011",
                "total": 999.99,
            },
            {
                "document_type": "invoice",
                "associated_bc_number": "16011",
                "subtotal": 900.00,
                "tps": 45.00,
                "tvq": 89.78,
                "total": 1034.78,
            },
        ])
        initial = {
            "bc_number": "17186",
            "subtotal": None,
            "tps": None,
            "tvq": None,
            "untaxed_extra_amount": None,
            "total": Decimal("54.58"),
        }
        _supplement_amounts_from_invoices(initial, receipt)
        self.assertEqual(initial["subtotal"], Decimal("47.47"))
        self.assertEqual(initial["tps"], Decimal("2.37"))
        self.assertEqual(initial["tvq"], Decimal("4.74"))

    def test_supplement_does_not_overwrite_existing_values(self):
        """When paper BC already has taxes, invoice values should NOT overwrite."""
        from bons.views import _supplement_amounts_from_invoices
        from unittest.mock import MagicMock

        receipt = MagicMock()
        receipt.ocr_raw_text = json.dumps([
            {
                "document_type": "paper_bc",
                "bc_number": "100",
                "subtotal": 10.00,
                "tps": 0.50,
                "tvq": 1.00,
                "total": 11.50,
            },
            {
                "document_type": "invoice",
                "subtotal": 99.99,
                "tps": 5.00,
                "tvq": 9.98,
                "total": 114.97,
            },
        ])
        initial = {
            "subtotal": Decimal("10.00"),
            "tps": Decimal("0.50"),
            "tvq": Decimal("1.00"),
            "untaxed_extra_amount": Decimal("2.00"),
            "total": Decimal("11.50"),
        }
        _supplement_amounts_from_invoices(initial, receipt)
        # Values should remain unchanged
        self.assertEqual(initial["subtotal"], Decimal("10.00"))
        self.assertEqual(initial["tps"], Decimal("0.50"))
        self.assertEqual(initial["tvq"], Decimal("1.00"))
        self.assertEqual(initial["untaxed_extra_amount"], Decimal("2.00"))
        self.assertEqual(initial["total"], Decimal("11.50"))

    def test_supplement_derives_standard_taxes_when_invoice_only_has_subtotal(self):
        from bons.views import _supplement_amounts_from_invoices
        from unittest.mock import MagicMock

        receipt = MagicMock()
        receipt.ocr_raw_text = json.dumps([
            {
                "document_type": "paper_bc",
                "bc_number": "16739",
                "subtotal": None,
                "tps": None,
                "tvq": None,
                "total": 19.49,
            },
            {
                "document_type": "invoice",
                "associated_bc_number": "16739",
                "subtotal": 16.95,
                "tps": None,
                "tvq": None,
                "total": None,
            },
        ])
        initial = {
            "subtotal": None,
            "tps": None,
            "tvq": None,
            "untaxed_extra_amount": None,
            "total": Decimal("19.49"),
        }
        _supplement_amounts_from_invoices(initial, receipt)
        self.assertEqual(initial["subtotal"], Decimal("16.95"))
        self.assertEqual(initial["tps"], Decimal("0.85"))
        self.assertEqual(initial["tvq"], Decimal("1.69"))

    def test_supplement_replaces_bc_subtotal_when_it_is_actually_total(self):
        from bons.views import _supplement_amounts_from_invoices
        from unittest.mock import MagicMock

        receipt = MagicMock()
        receipt.ocr_raw_text = json.dumps([
            {
                "document_type": "paper_bc",
                "bc_number": "16739",
                "subtotal": 19.49,
                "tps": None,
                "tvq": None,
                "total": 19.49,
            },
            {
                "document_type": "invoice",
                "associated_bc_number": "16739",
                "subtotal": 16.95,
                "tps": None,
                "tvq": None,
                "total": 16.95,
            },
        ])
        initial = {
            "subtotal": Decimal("19.49"),
            "tps": None,
            "tvq": None,
            "untaxed_extra_amount": None,
            "total": Decimal("19.49"),
        }
        _supplement_amounts_from_invoices(initial, receipt)
        self.assertEqual(initial["subtotal"], Decimal("16.95"))
        self.assertEqual(initial["tps"], Decimal("0.85"))
        self.assertEqual(initial["tvq"], Decimal("1.69"))


class SignerRoleWorkflowTests(TestCase):
    def setUp(self):
        self.house = House.objects.create(code="BB", name="Maison BB", account_number="13-51200")
        self.budget_year = BudgetYear.objects.create(
            house=self.house, year=2026, annual_budget_total=Decimal("12237.00")
        )
        self.sub_budget = SubBudget.objects.create(
            budget_year=self.budget_year, trace_code=7,
            name="Produits ménager", planned_amount=Decimal("300.00")
        )
        self.purchaser = Member.objects.create(first_name="Marylin", last_name="Lamarche")
        self.validator = Member.objects.create(first_name="René", last_name="Côté")
        self.treasurer_member = Member.objects.create(first_name="Trésorier", last_name="Test")
        self.apt_202 = Apartment.objects.create(house=self.house, code="202")
        self.apt_203 = Apartment.objects.create(house=self.house, code="203")
        self.apt_204 = Apartment.objects.create(house=self.house, code="204")
        Residency.objects.create(member=self.purchaser, apartment=self.apt_202, start_date=date(2020, 1, 1))
        Residency.objects.create(member=self.validator, apartment=self.apt_203, start_date=date(2020, 1, 1))
        Residency.objects.create(member=self.treasurer_member, apartment=self.apt_204, start_date=date(2020, 1, 1))
        self.user = User.objects.create_user(
            username="tresorier",
            password="test123",
            role=User.Role.TREASURER,
            house=self.house,
            member=self.treasurer_member,
        )
        self.bon = BonDeCommande.objects.create(
            house=self.house,
            budget_year=self.budget_year,
            number="16011",
            purchase_date=date(2026, 1, 7),
            short_description="Paper BC",
            total=Decimal("547.93"),
            sub_budget=self.sub_budget,
            purchaser_member=self.treasurer_member,
            status=BonStatus.READY_FOR_VALIDATION,
            is_paper_bc=True,
            paper_bc_number="16011",
        )
        self.receipt = ReceiptFile.objects.create(
            bon_de_commande=self.bon,
            file=SimpleUploadedFile("BC16011.pdf", b"fake pdf", content_type="application/pdf"),
            original_filename="BC16011.pdf",
            content_type="application/pdf",
            ocr_status="EXTRACTED",
        )
        self.extracted = ReceiptExtractedFields.objects.create(
            receipt_file=self.receipt,
            document_type_candidate="paper_bc",
            final_document_type="paper_bc",
            bc_number_candidate="16011",
            final_bc_number="16011",
            expense_member_name_candidate="Marylin Lamarche",
            expense_apartment_candidate="202",
            validator_member_name_candidate="René Côté",
            validator_apartment_candidate="203",
            signer_roles_ambiguous_candidate=True,
            total_candidate=Decimal("547.93"),
            final_total=Decimal("547.93"),
            sub_budget=self.sub_budget,
        )

    def test_receipt_review_persists_swapped_signers_for_existing_bon(self):
        self.client.login(username="tresorier", password="test123")
        prefix = f"receipt_{self.receipt.pk}"
        response = self.client.post(
            reverse("bons:receipt-review", kwargs={"bon_pk": self.bon.pk, "receipt_pk": self.receipt.pk}),
            {
                f"{prefix}-document_type": "paper_bc",
                f"{prefix}-bc_number": "16011",
                f"{prefix}-associated_bc_number": "",
                f"{prefix}-supplier_name": "Gicleurs",
                f"{prefix}-supplier_address": "",
                f"{prefix}-expense_member_name": "René Côté",
                f"{prefix}-expense_apartment": "203",
                f"{prefix}-expense_member": self.validator.pk,
                f"{prefix}-validator_member_name": "Marylin Lamarche",
                f"{prefix}-validator_apartment": "202",
                f"{prefix}-validator_member": self.purchaser.pk,
                f"{prefix}-signer_roles_ambiguous": "on",
                f"{prefix}-member_name_raw": "",
                f"{prefix}-apartment_number": "",
                f"{prefix}-purchaser_member": "",
                f"{prefix}-matched_member_id": "",
                f"{prefix}-sub_budget": self.sub_budget.pk,
                f"{prefix}-merchant_name": "",
                f"{prefix}-purchase_date": "2026-01-07",
                f"{prefix}-subtotal": "",
                f"{prefix}-tps": "",
                f"{prefix}-tvq": "",
                f"{prefix}-total": "547.93",
                f"{prefix}-summary": "Travaux",
            },
        )
        self.assertEqual(response.status_code, 302)
        self.extracted.refresh_from_db()
        self.assertEqual(self.extracted.final_expense_member_name, "René Côté")
        self.assertEqual(self.extracted.final_validator_member_name, "Marylin Lamarche")
        self.assertTrue(self.extracted.signer_roles_ambiguous_final)

        self.bon.refresh_from_db()
        self.assertEqual(self.bon.purchaser_member, self.validator)
        self.assertEqual(self.bon.approver_member, self.purchaser)

    def test_build_form_prefills_paper_bc_merchant_from_supplier(self):
        from bons.views import OcrReviewView

        self.extracted.final_supplier_name = "Produits Sany"
        self.extracted.final_merchant = ""
        self.extracted.save()

        form, *_ = OcrReviewView()._build_form(self.receipt, self.bon)
        self.assertEqual(form.initial["merchant_name"], "Produits Sany")

    def test_build_form_prefers_invoice_supplier_name_and_address_for_company_reimbursement(self):
        from bons.views import OcrReviewView

        self.extracted.final_supplier_name = "Parent"
        self.extracted.final_supplier_address = "Rock Forest"
        self.extracted.final_merchant = ""
        self.receipt.ocr_raw_text = json.dumps([
            {
                "document_type": "paper_bc",
                "supplier_name": "Parent",
                "supplier_address": "Rock Forest",
            },
            {
                "document_type": "invoice",
                "supplier_name": "Quincaillerie Parent Enr.",
                "supplier_address": "1237 Belvedere Sud, Sherbrooke, QC, J1H 4E1",
            },
        ])
        self.receipt.save(update_fields=["ocr_raw_text"])
        self.extracted.save()

        form, *_ = OcrReviewView()._build_form(self.receipt, self.bon)
        self.assertEqual(form.initial["supplier_name"], "Quincaillerie Parent Enr.")
        self.assertEqual(
            form.initial["supplier_address"],
            "1237 Belvedere Sud, Sherbrooke, QC, J1H 4E1",
        )
        self.assertEqual(form.initial["merchant_name"], "Quincaillerie Parent Enr.")
        self.assertEqual(form.initial["reimburse_to"], "supplier")

    def test_build_form_keeps_member_reimbursement_from_paper_bc_payee(self):
        from bons.views import OcrReviewView

        self.extracted.final_supplier_name = "Marylin Lamarche"
        self.extracted.supplier_name_candidate = "Marylin Lamarche"
        self.receipt.ocr_raw_text = json.dumps([
            {
                "document_type": "paper_bc",
                "supplier_name": "Marylin Lamarche",
                "expense_member_name": "Marylin Lamarche",
            },
            {
                "document_type": "invoice",
                "supplier_name": "Quincaillerie Parent Enr.",
                "supplier_address": "1237 Belvedere Sud, Sherbrooke, QC, J1H 4E1",
            },
        ])
        self.receipt.save(update_fields=["ocr_raw_text"])
        self.extracted.save()

        form, *_ = OcrReviewView()._build_form(self.receipt, self.bon)
        self.assertEqual(form.initial["supplier_name"], "Quincaillerie Parent Enr.")
        self.assertEqual(form.initial["reimburse_to"], "member")

    def test_confidence_summary_prefers_confirmed_false_signer_ambiguity(self):
        from bons.ai_confidence import build_receipt_confidence_summary_rows

        self.extracted.signer_roles_ambiguous_candidate = True
        self.extracted.signer_roles_ambiguous_final = False
        self.extracted.final_confidence_scores = {"signer_roles_ambiguous": 4}
        self.extracted.save(update_fields=[
            "signer_roles_ambiguous_candidate",
            "signer_roles_ambiguous_final",
            "final_confidence_scores",
        ])

        rows = build_receipt_confidence_summary_rows(self.receipt)
        summary = next(row for row in rows if row["field_name"] == "signer_roles_ambiguous")
        self.assertEqual(summary["value"], "Non")

    def test_receipt_review_renders_mobile_layout_hooks(self):
        self.client.login(username="tresorier", password="test123")

        response = self.client.get(
            reverse(
                "bons:receipt-review",
                kwargs={"bon_pk": self.bon.pk, "receipt_pk": self.receipt.pk},
            )
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'class="review-layout"')
        self.assertContains(response, 'class="review-navigation review-navigation-single"')

    def test_build_review_confidence_marks_corrected_fields_as_na(self):
        from bons.ai_confidence import build_receipt_review_confidence_scores

        self.receipt.ocr_raw_text = json.dumps([
            {
                "document_type": "paper_bc",
                "supplier_name": "Fournisseur OCR",
                "field_confidence_scores": {
                    "supplier_name": 8,
                },
            }
        ])
        self.receipt.save(update_fields=["ocr_raw_text"])

        scores = build_receipt_review_confidence_scores(
            self.receipt,
            {"supplier_name": "Fournisseur corrigé"},
            document_type="paper_bc",
        )

        self.assertEqual(scores["supplier_name"], "NA")

    def test_receipt_review_syncs_member_reimbursement_from_paper_bc_payee(self):
        self.client.login(username="tresorier", password="test123")
        prefix = f"receipt_{self.receipt.pk}"
        self.extracted.supplier_name_candidate = "Marylin Lamarche"
        self.receipt.ocr_raw_text = json.dumps([
            {
                "document_type": "paper_bc",
                "supplier_name": "Marylin Lamarche",
                "expense_member_name": "Marylin Lamarche",
            },
            {
                "document_type": "invoice",
                "supplier_name": "Quincaillerie Parent Enr.",
                "supplier_address": "1237 Belvedere Sud, Sherbrooke, QC, J1H 4E1",
            },
        ])
        self.receipt.save(update_fields=["ocr_raw_text"])
        self.extracted.save(update_fields=["supplier_name_candidate"])

        response = self.client.post(
            reverse("bons:receipt-review", kwargs={"bon_pk": self.bon.pk, "receipt_pk": self.receipt.pk}),
            {
                f"{prefix}-document_type": "paper_bc",
                f"{prefix}-bc_number": "16011",
                f"{prefix}-associated_bc_number": "",
                f"{prefix}-supplier_name": "Quincaillerie Parent Enr.",
                f"{prefix}-supplier_address": "1237 Belvedere Sud, Sherbrooke, QC, J1H 4E1",
                f"{prefix}-reimburse_to": "supplier",
                f"{prefix}-expense_member_name": "Marylin Lamarche",
                f"{prefix}-expense_apartment": "202",
                f"{prefix}-expense_member": self.purchaser.pk,
                f"{prefix}-validator_member_name": "René Côté",
                f"{prefix}-validator_apartment": "203",
                f"{prefix}-validator_member": self.validator.pk,
                f"{prefix}-signer_roles_ambiguous": "on",
                f"{prefix}-member_name_raw": "",
                f"{prefix}-apartment_number": "",
                f"{prefix}-purchaser_member": "",
                f"{prefix}-matched_member_id": "",
                f"{prefix}-sub_budget": self.sub_budget.pk,
                f"{prefix}-merchant_name": "Quincaillerie Parent Enr.",
                f"{prefix}-purchase_date": "2026-01-07",
                f"{prefix}-subtotal": "",
                f"{prefix}-tps": "",
                f"{prefix}-tvq": "",
                f"{prefix}-total": "547.93",
                f"{prefix}-summary": "Travaux",
            },
        )
        self.assertEqual(response.status_code, 302)

        self.extracted.refresh_from_db()
        self.bon.refresh_from_db()
        self.assertEqual(self.extracted.final_reimburse_to, "member")
        self.assertEqual(self.bon.reimburse_to, "member")

    def test_receipt_review_derives_taxes_with_untaxed_extra(self):
        self.client.login(username="tresorier", password="test123")
        prefix = f"receipt_{self.receipt.pk}"
        response = self.client.post(
            reverse("bons:receipt-review", kwargs={"bon_pk": self.bon.pk, "receipt_pk": self.receipt.pk}),
            {
                f"{prefix}-document_type": "paper_bc",
                f"{prefix}-bc_number": "16011",
                f"{prefix}-associated_bc_number": "",
                f"{prefix}-supplier_name": "Produits Sany",
                f"{prefix}-supplier_address": "",
                f"{prefix}-expense_member_name": "Marylin Lamarche",
                f"{prefix}-expense_apartment": "202",
                f"{prefix}-expense_member": self.purchaser.pk,
                f"{prefix}-validator_member_name": "René Côté",
                f"{prefix}-validator_apartment": "203",
                f"{prefix}-validator_member": self.validator.pk,
                f"{prefix}-signer_roles_ambiguous": "on",
                f"{prefix}-member_name_raw": "",
                f"{prefix}-apartment_number": "",
                f"{prefix}-purchaser_member": "",
                f"{prefix}-matched_member_id": "",
                f"{prefix}-sub_budget": self.sub_budget.pk,
                f"{prefix}-merchant_name": "",
                f"{prefix}-purchase_date": "2026-01-07",
                f"{prefix}-subtotal": "100.00",
                f"{prefix}-tps": "",
                f"{prefix}-tvq": "",
                f"{prefix}-untaxed_extra_amount": "10.00",
                f"{prefix}-total": "124.98",
                f"{prefix}-summary": "Travaux",
            },
        )
        self.assertEqual(response.status_code, 302)
        self.extracted.refresh_from_db()
        self.assertEqual(self.extracted.final_subtotal, Decimal("100.00"))
        self.assertEqual(self.extracted.final_tps, Decimal("5.00"))
        self.assertEqual(self.extracted.final_tvq, Decimal("9.98"))
        self.assertEqual(self.extracted.final_untaxed_extra_amount, Decimal("10.00"))
        self.assertEqual(self.extracted.final_total, Decimal("124.98"))

    def test_swap_signers_view_swaps_existing_bon(self):
        self.bon.purchaser_member = self.purchaser
        self.bon.purchaser_apartment = self.apt_202
        self.bon.approver_member = self.validator
        self.bon.approver_apartment = self.apt_203
        self.bon.save()

        self.client.login(username="tresorier", password="test123")
        response = self.client.post(reverse("bons:swap-signers", kwargs={"pk": self.bon.pk}))
        self.assertEqual(response.status_code, 302)
        self.bon.refresh_from_db()
        self.assertEqual(self.bon.purchaser_member, self.validator)
        self.assertEqual(self.bon.purchaser_apartment, self.apt_203)
        self.assertEqual(self.bon.approver_member, self.purchaser)
        self.assertEqual(self.bon.approver_apartment, self.apt_202)

    def test_external_supplier_signer_no_mismatch(self):
        """When the validator name doesn't match any member, treat as external (no mismatch)."""
        self.extracted.validator_member_name_candidate = "Plomberie ABC Inc."
        self.extracted.validator_apartment_candidate = ""
        self.extracted.save()

        initial, _, validator_mismatch, _, _ = _paper_bc_signer_initials(self.house, self.extracted)
        self.assertFalse(validator_mismatch)
        self.assertTrue(initial.get("validator_is_external"))

    def test_external_validator_name_overrides_bad_apartment_guess(self):
        """A conflicting apartment guess must not force the validator to a house member."""
        extra_member = Member.objects.create(first_name="Lionel", last_name="Munezero")
        apt_104 = Apartment.objects.create(house=self.house, code="104")
        Residency.objects.create(member=extra_member, apartment=apt_104, start_date=date(2020, 1, 1))
        self.extracted.validator_member_name_candidate = "Lui-Jade Rodrigue"
        self.extracted.validator_apartment_candidate = "104"
        self.extracted.save()

        initial, _, validator_mismatch, _, _ = _paper_bc_signer_initials(self.house, self.extracted)
        self.assertFalse(validator_mismatch)
        self.assertTrue(initial.get("validator_is_external"))
        self.assertEqual(initial.get("validator_member_name"), "Lui-Jade Rodrigue")
        self.assertEqual(initial.get("validator_apartment"), "")
        self.assertNotIn("validator_member", initial)

    def test_confirmed_false_signer_ambiguity_overrides_candidate_true(self):
        self.extracted.signer_roles_ambiguous_candidate = True
        self.extracted.signer_roles_ambiguous_final = False
        self.extracted.confirmed_at = timezone.now()
        self.extracted.save(update_fields=[
            "signer_roles_ambiguous_candidate",
            "signer_roles_ambiguous_final",
            "confirmed_at",
        ])

        initial, _, _, _, _ = _paper_bc_signer_initials(self.house, self.extracted)
        self.assertFalse(initial["signer_roles_ambiguous"])

    def test_external_supplier_display_label(self):
        """External approver should display as 'FOUR / [name]'."""
        self.bon.approver_is_external = True
        self.bon.approver_member = None
        self.bon.approver_apartment = None
        self.bon.approver_name_snapshot = "Plomberie ABC Inc."
        self.bon.save()

        self.assertEqual(self.bon.approver_display_label, "FOUR / Plomberie ABC Inc.")
        self.assertEqual(self.bon.effective_validator_display_label, "FOUR / Plomberie ABC Inc.")

    def test_sync_existing_bon_keeps_external_validator_when_apartment_conflicts(self):
        from bons.views import OcrReviewView

        extra_member = Member.objects.create(first_name="Lionel", last_name="Munezero")
        apt_104 = Apartment.objects.create(house=self.house, code="104")
        Residency.objects.create(member=extra_member, apartment=apt_104, start_date=date(2020, 1, 1))
        self.extracted.validator_member_name_candidate = "Lui-Jade Rodrigue"
        self.extracted.validator_apartment_candidate = "104"
        self.extracted.save()

        OcrReviewView()._sync_existing_bon_paper_bc_data(self.bon)
        self.bon.refresh_from_db()

        self.assertTrue(self.bon.approver_is_external)
        self.assertIsNone(self.bon.approver_member)
        self.assertIsNone(self.bon.approver_apartment)
        self.assertEqual(self.bon.approver_name_snapshot, "Lui-Jade Rodrigue")

    def test_external_supplier_blocks_swap(self):
        """Swapping signers should be blocked for external approvers."""
        self.bon.approver_is_external = True
        self.bon.approver_member = None
        self.bon.approver_apartment = None
        self.bon.approver_name_snapshot = "Plomberie ABC Inc."
        self.bon.save()

        self.client.login(username="tresorier", password="test123")
        response = self.client.post(reverse("bons:swap-signers", kwargs={"pk": self.bon.pk}))
        self.assertEqual(response.status_code, 302)
        self.bon.refresh_from_db()
        # Purchaser should NOT have changed
        self.assertEqual(self.bon.purchaser_member, self.treasurer_member)

    def test_external_snapshot_preserved_on_refresh(self):
        """refresh_snapshot_fields should keep the external name."""
        self.bon.approver_is_external = True
        self.bon.approver_member = None
        self.bon.approver_apartment = None
        self.bon.approver_name_snapshot = "Plomberie ABC Inc."
        self.bon.refresh_snapshot_fields()
        self.assertEqual(self.bon.approver_name_snapshot, "Plomberie ABC Inc.")


class FuzzyNameMatchTests(TestCase):
    """Test _names_match and _normalize_name for case/accent-insensitive matching."""

    def test_exact_match(self):
        self.assertTrue(_names_match("Marylin Lamarche", "Marylin Lamarche"))

    def test_case_insensitive(self):
        self.assertTrue(_names_match("MARYLINE LAMARCHE", "Marylin Lamarche"))

    def test_spelling_variation(self):
        # MARYLINE vs Marylin — one extra letter
        self.assertTrue(_names_match("MARYLINE LAMARCHE", "Marylin Lamarche"))

    def test_accented_names(self):
        self.assertTrue(_names_match("René Lévesque", "rene levesque"))
        self.assertTrue(_names_match("HÉLOÏSE CÔTÉ", "Heloise Cote"))

    def test_substring_match(self):
        self.assertTrue(_names_match("Carl-David", "Carl-David Fortin"))

    def test_completely_different_names(self):
        self.assertFalse(_names_match("Jean Tremblay", "Marie Bouchard"))

    def test_empty_strings(self):
        self.assertFalse(_names_match("", "Someone"))
        self.assertFalse(_names_match("Someone", ""))
        self.assertFalse(_names_match("", ""))

    def test_normalize_strips_accents(self):
        self.assertEqual(_normalize_name("Héloïse Côté"), "heloise cote")
        self.assertEqual(_normalize_name("  CARL-DAVID  FORTIN  "), "carl-david fortin")


# ═══════════════════════════════════════════════════════════════════════════
# Phase 10 Tests: Duplicate Detection, Export Gating, Audit Trail
# ═══════════════════════════════════════════════════════════════════════════

class DuplicateDetectionBaseTest(TestCase):
    """Base setup for duplicate detection tests."""

    def setUp(self):
        from django.core.files.uploadedfile import SimpleUploadedFile
        from bons.models import ReceiptFile, ReceiptExtractedFields, OcrStatus

        self.house = House.objects.create(
            code="BB", name="Maison BB", account_number="13-51200"
        )
        self.budget_year = BudgetYear.objects.create(
            house=self.house, year=2026,
            annual_budget_total=Decimal("12237.00")
        )
        self.sub_budget = SubBudget.objects.create(
            budget_year=self.budget_year, trace_code=7,
            name="Produits ménager", planned_amount=Decimal("300.00")
        )
        self.member = Member.objects.create(first_name="Marylin", last_name="Lamarche")
        self.apartment = Apartment.objects.create(house=self.house, code="202")
        Residency.objects.create(
            member=self.member, apartment=self.apartment,
            start_date=date(2020, 1, 1)
        )
        self.user = User.objects.create_user(
            username="tresorier", password="test123",
            role="TREASURER", house=self.house, member=self.member,
        )

        # Create existing bon with a receipt
        self.existing_bon = BonDeCommande.objects.create(
            house=self.house, budget_year=self.budget_year,
            number="BB260001", purchase_date=date(2026, 1, 7),
            short_description="Existing purchase", total=Decimal("19.49"),
            sub_budget=self.sub_budget, purchaser_member=self.member,
            status=BonStatus.VALIDATED,
        )
        fake_file = SimpleUploadedFile("receipt1.png", b"fake image data", content_type="image/png")
        self.existing_receipt = ReceiptFile.objects.create(
            bon_de_commande=self.existing_bon,
            file=fake_file,
            original_filename="receipt1.png",
            content_type="image/png",
            ocr_status=OcrStatus.EXTRACTED,
        )
        self.existing_ef = ReceiptExtractedFields.objects.create(
            receipt_file=self.existing_receipt,
            total_candidate=Decimal("19.49"),
            final_total=Decimal("19.49"),
        )

        # Create new bon with a receipt (same total)
        self.new_bon = BonDeCommande.objects.create(
            house=self.house, budget_year=self.budget_year,
            number="BB260002", purchase_date=date(2026, 3, 15),
            short_description="New purchase", total=Decimal("19.49"),
            sub_budget=self.sub_budget, purchaser_member=self.member,
            status=BonStatus.READY_FOR_VALIDATION,
        )
        fake_file2 = SimpleUploadedFile("receipt2.png", b"fake image data 2", content_type="image/png")
        self.new_receipt = ReceiptFile.objects.create(
            bon_de_commande=self.new_bon,
            file=fake_file2,
            original_filename="receipt2.png",
            content_type="image/png",
            ocr_status=OcrStatus.EXTRACTED,
        )
        self.new_ef = ReceiptExtractedFields.objects.create(
            receipt_file=self.new_receipt,
            total_candidate=Decimal("19.49"),
            final_total=Decimal("19.49"),
        )


class DigitalInvoiceDuplicateTests(DuplicateDetectionBaseTest):
    """Test digital invoice duplicate detection via matching totals."""

    def test_find_matching_totals_same_total(self):
        matches = DuplicateDetectionService.find_matching_totals(
            self.new_receipt, self.house
        )
        self.assertEqual(len(matches), 1)
        self.assertEqual(matches[0].receipt_file, self.existing_receipt)

    def test_find_matching_totals_different_total(self):
        self.existing_ef.final_total = Decimal("99.99")
        self.existing_ef.total_candidate = Decimal("99.99")
        self.existing_ef.save()
        matches = DuplicateDetectionService.find_matching_totals(
            self.new_receipt, self.house
        )
        self.assertEqual(len(matches), 0)

    def test_find_matching_totals_different_house(self):
        other_house = House.objects.create(code="XX", name="Other", account_number="99-99999")
        matches = DuplicateDetectionService.find_matching_totals(
            self.new_receipt, other_house
        )
        self.assertEqual(len(matches), 0)

    def test_find_matching_totals_ignores_scan_session_receipts(self):
        temp_bon = BonDeCommande.objects.create(
            house=self.house,
            budget_year=self.budget_year,
            number="BB260099",
            purchase_date=date(2026, 2, 1),
            short_description="Temp scan",
            total=Decimal("19.49"),
            sub_budget=self.sub_budget,
            purchaser_member=self.member,
            status=BonStatus.READY_FOR_REVIEW,
            is_scan_session=True,
        )
        temp_receipt = ReceiptFile.objects.create(
            bon_de_commande=temp_bon,
            file=SimpleUploadedFile("temp.png", b"fake temp image", content_type="image/png"),
            original_filename="temp.png",
            content_type="image/png",
            ocr_status=OcrStatus.EXTRACTED,
        )
        ReceiptExtractedFields.objects.create(
            receipt_file=temp_receipt,
            total_candidate=Decimal("19.49"),
            final_total=Decimal("19.49"),
        )

        matches = DuplicateDetectionService.find_matching_totals(
            self.new_receipt, self.house
        )

        self.assertEqual(len(matches), 1)
        self.assertEqual(matches[0].receipt_file, self.existing_receipt)

    def test_find_matching_totals_ignores_voided_bons(self):
        self.existing_bon.status = BonStatus.VOID
        self.existing_bon.save()

        matches = DuplicateDetectionService.find_matching_totals(
            self.new_receipt, self.house
        )
        self.assertEqual(len(matches), 0)

    def test_find_matching_totals_ignores_archived_receipts(self):
        self.existing_receipt.archive("Archived during bon reactivation")

        matches = DuplicateDetectionService.find_matching_totals(
            self.new_receipt, self.house
        )
        self.assertEqual(len(matches), 0)

    def test_normalize_confidence_accepts_percent_values(self):
        self.assertEqual(
            DuplicateDetectionService._normalize_confidence("99%"),
            0.99,
        )
        self.assertEqual(
            DuplicateDetectionService._normalize_confidence(95),
            0.95,
        )

    @patch.object(DuplicateDetectionService, "_receipt_to_base64", return_value="ZmFrZQ==")
    @patch("bons.ocr_service.OpenAI")
    def test_compare_with_gpt_treats_false_string_as_false(
        self,
        mock_openai,
        mock_receipt_to_base64,
    ):
        response = MagicMock()
        response.choices = [
            MagicMock(
                message=MagicMock(
                    content=json.dumps(
                        {
                            "is_same_purchase": "false",
                            "confidence": 0.82,
                            "reasoning": "Transactions differentes",
                            "field_confidence_scores": {
                                "is_same_purchase": 2,
                                "confidence": 8,
                                "reasoning": 7,
                            },
                        }
                    )
                )
            )
        ]
        mock_openai.return_value.chat.completions.create.return_value = response

        with self.settings(OPENAI_API_KEY="test-key"):
            result = DuplicateDetectionService.compare_with_gpt(
                self.new_receipt,
                self.existing_receipt,
            )

        self.assertFalse(result["is_same_purchase"])
        self.assertEqual(result["field_confidence_scores"]["is_same_purchase"], 2)

    @patch.object(DuplicateDetectionService, "compare_with_gpt")
    def test_check_and_flag_creates_flag(self, mock_gpt):
        mock_gpt.return_value = {
            "is_same_purchase": True,
            "confidence": 0.95,
            "reasoning": "Same merchant, date, and amount",
            "field_confidence_scores": {
                "is_same_purchase": 9,
                "confidence": 8,
                "reasoning": 7,
            },
        }
        flags = DuplicateDetectionService.check_and_flag_duplicates(
            self.new_receipt, self.house
        )
        self.assertEqual(len(flags), 1)
        flag = flags[0]
        self.assertEqual(flag.receipt_file, self.new_receipt)
        self.assertEqual(flag.suspected_duplicate_receipt, self.existing_receipt)
        self.assertEqual(flag.confidence, Decimal("0.95"))
        self.assertEqual(flag.status, DuplicateFlagStatus.CONFIRMED_DUPLICATE)
        self.assertEqual(flag.field_confidence_scores["is_same_purchase"], 9)
        self.assertEqual(flag.field_confidence_scores["reasoning"], 7)

    @patch.object(DuplicateDetectionService, "compare_with_gpt")
    def test_check_and_flag_pending_if_low_confidence(self, mock_gpt):
        mock_gpt.return_value = {
            "is_same_purchase": True,
            "confidence": 0.60,
            "reasoning": "Similar amounts but different dates",
        }
        flags = DuplicateDetectionService.check_and_flag_duplicates(
            self.new_receipt, self.house
        )
        self.assertEqual(len(flags), 1)
        self.assertEqual(flags[0].status, DuplicateFlagStatus.PENDING)

    @patch.object(DuplicateDetectionService, "compare_with_gpt")
    def test_no_duplicate_flag_for_same_receipt(self, mock_gpt):
        """Should not flag a receipt against itself."""
        mock_gpt.return_value = {
            "is_same_purchase": True, "confidence": 1.0, "reasoning": "Same"
        }
        matches = DuplicateDetectionService.find_matching_totals(
            self.existing_receipt, self.house
        )
        # The existing receipt should match the new one, not itself
        self.assertNotIn(
            self.existing_receipt,
            [m.receipt_file for m in matches]
        )

    @patch.object(DuplicateDetectionService, "compare_with_gpt")
    def test_no_double_flagging(self, mock_gpt):
        mock_gpt.return_value = {
            "is_same_purchase": True, "confidence": 0.95, "reasoning": "Dup"
        }
        # First call creates flags
        flags1 = DuplicateDetectionService.check_and_flag_duplicates(
            self.new_receipt, self.house
        )
        self.assertEqual(len(flags1), 1)

        # Second call should not create duplicates
        flags2 = DuplicateDetectionService.check_and_flag_duplicates(
            self.new_receipt, self.house
        )
        self.assertEqual(len(flags2), 0)


class PaperBcDuplicateTests(DuplicateDetectionBaseTest):
    """Test paper BC duplicate detection at finalization."""

    def test_paper_bc_same_number_same_total_blocks(self):
        """If a paper BC with the same number and total exists, it should be blocked."""
        self.existing_bon.is_paper_bc = True
        self.existing_bon.paper_bc_number = "16011"
        BonDeCommande.objects.filter(pk=self.existing_bon.pk).update(
            is_paper_bc=True, paper_bc_number="16011"
        )

        existing_dup = BonDeCommande.objects.filter(
            house=self.house,
            is_paper_bc=True,
            paper_bc_number="16011",
        ).exclude(status=BonStatus.VOID).first()

        self.assertIsNotNone(existing_dup)
        self.assertEqual(existing_dup.total, Decimal("19.49"))

    def test_paper_bc_voided_not_detected(self):
        """Voided paper BCs should not block new ones."""
        BonDeCommande.objects.filter(pk=self.existing_bon.pk).update(
            is_paper_bc=True, paper_bc_number="16011", status=BonStatus.VOID,
        )

        existing_dup = BonDeCommande.objects.filter(
            house=self.house,
            is_paper_bc=True,
            paper_bc_number="16011",
        ).exclude(status=BonStatus.VOID).first()

        self.assertIsNone(existing_dup)


class ExportGatingTests(DuplicateDetectionBaseTest):
    """Test that exports are blocked when receipts haven't been reviewed."""

    def test_export_ready_when_all_reviewed(self):
        self.existing_receipt.ocr_status = OcrStatus.CORRECTED
        self.existing_receipt.save(update_fields=["ocr_status"])
        self.assertTrue(_bon_is_export_ready(self.existing_bon))

    def test_export_not_ready_when_only_extracted(self):
        self.assertFalse(_bon_is_export_ready(self.existing_bon))

    def test_export_not_ready_when_pending(self):
        from bons.models import ReceiptFile, OcrStatus
        from django.core.files.uploadedfile import SimpleUploadedFile
        fake = SimpleUploadedFile("unreviewed.png", b"data", content_type="image/png")
        ReceiptFile.objects.create(
            bon_de_commande=self.existing_bon,
            file=fake, original_filename="unreviewed.png",
            content_type="image/png",
            ocr_status=OcrStatus.PENDING,
        )
        self.assertFalse(_bon_is_export_ready(self.existing_bon))

    def test_export_ready_ignores_archived_pending_receipt(self):
        self.existing_receipt.ocr_status = OcrStatus.CORRECTED
        self.existing_receipt.save(update_fields=["ocr_status"])
        fake = SimpleUploadedFile("archived-pending.png", b"data", content_type="image/png")
        receipt = ReceiptFile.objects.create(
            bon_de_commande=self.existing_bon,
            file=fake,
            original_filename="archived-pending.png",
            content_type="image/png",
            ocr_status=OcrStatus.PENDING,
        )
        receipt.archive("Archived during bon reactivation")
        self.assertTrue(_bon_is_export_ready(self.existing_bon))

    def test_export_ready_no_receipts(self):
        empty_bon = BonDeCommande.objects.create(
            house=self.house, budget_year=self.budget_year,
            number="BB260099", purchase_date=date(2026, 1, 1),
            short_description="Empty bon", total=Decimal("0.00"),
            sub_budget=self.sub_budget, purchaser_member=self.member,
        )
        self.assertTrue(_bon_is_export_ready(empty_bon))

    def test_export_pdf_blocked_when_not_ready(self):
        from bons.models import ReceiptFile, OcrStatus
        from django.core.files.uploadedfile import SimpleUploadedFile
        fake = SimpleUploadedFile("pending.png", b"data", content_type="image/png")
        ReceiptFile.objects.create(
            bon_de_commande=self.existing_bon,
            file=fake, original_filename="pending.png",
            content_type="image/png",
            ocr_status=OcrStatus.NOT_REQUESTED,
        )
        self.client.login(username="tresorier", password="test123")
        resp = self.client.get(f"/bons/{self.existing_bon.pk}/pdf/")
        self.assertEqual(resp.status_code, 302)  # redirect back to detail

    def test_export_xlsx_blocked_when_not_ready(self):
        from bons.models import ReceiptFile, OcrStatus
        from django.core.files.uploadedfile import SimpleUploadedFile
        fake = SimpleUploadedFile("pending.png", b"data", content_type="image/png")
        ReceiptFile.objects.create(
            bon_de_commande=self.existing_bon,
            file=fake, original_filename="pending.png",
            content_type="image/png",
            ocr_status=OcrStatus.NOT_REQUESTED,
        )
        self.client.login(username="tresorier", password="test123")
        resp = self.client.get(f"/bons/{self.existing_bon.pk}/xlsx/")
        self.assertEqual(resp.status_code, 302)  # redirect back to detail


class ExportDuplicateVisualAttachmentTests(DuplicateDetectionBaseTest):
    def setUp(self):
        super().setUp()
        import io
        from PIL import Image

        def png_bytes(color):
            buf = io.BytesIO()
            Image.new("RGB", (80, 80), color).save(buf, format="PNG")
            return buf.getvalue()

        self.existing_receipt.archive("Replaced with valid export preview image")
        self.new_receipt.archive("Replaced with valid export preview image")

        self.existing_receipt = ReceiptFile.objects.create(
            bon_de_commande=self.existing_bon,
            file=SimpleUploadedFile("receipt1.png", png_bytes((255, 0, 0)), content_type="image/png"),
            original_filename="receipt1.png",
            content_type="image/png",
            ocr_status=OcrStatus.EXTRACTED,
        )
        ReceiptExtractedFields.objects.create(
            receipt_file=self.existing_receipt,
            total_candidate=Decimal("19.49"),
            final_total=Decimal("19.49"),
        )

        self.new_receipt = ReceiptFile.objects.create(
            bon_de_commande=self.new_bon,
            file=SimpleUploadedFile("receipt2.png", png_bytes((0, 0, 255)), content_type="image/png"),
            original_filename="receipt2.png",
            content_type="image/png",
            ocr_status=OcrStatus.EXTRACTED,
        )
        ReceiptExtractedFields.objects.create(
            receipt_file=self.new_receipt,
            total_candidate=Decimal("19.49"),
            final_total=Decimal("19.49"),
        )

        self.flag = DuplicateFlag.objects.create(
            receipt_file=self.new_receipt,
            suspected_duplicate_receipt=self.existing_receipt,
            confidence=Decimal("0.95"),
            gpt_comparison_result="Même achat probable",
            status=DuplicateFlagStatus.PENDING,
        )

    def test_generate_bon_pdf_includes_duplicate_warning_and_extra_pages(self):
        from bons.pdf_service import generate_bon_pdf

        pdf_bytes = generate_bon_pdf(self.new_bon)
        self.assertGreaterEqual(pdf_bytes.count(b"/Type /Page"), 3)
        self.assertGreaterEqual(pdf_bytes.count(b"/Subtype /Image"), 2)

    def test_generate_bon_xlsx_creates_duplicate_sheet_with_images(self):
        import io
        import zipfile
        from openpyxl import load_workbook
        from bons.pdf_service import generate_bon_xlsx

        xlsx_bytes = generate_bon_xlsx(self.new_bon)
        workbook = load_workbook(io.BytesIO(xlsx_bytes))
        self.assertIn("Doublons possibles", workbook.sheetnames)
        duplicate_sheet = workbook["Doublons possibles"]
        values = [
            str(value)
            for row in duplicate_sheet.iter_rows(values_only=True)
            for value in row
            if value is not None
        ]
        joined_values = " ".join(values)
        self.assertIn("POSSIBLE DOUBLON", joined_values)
        self.assertIn("receipt1.png", joined_values)
        self.assertIn("receipt2.png", joined_values)

        with zipfile.ZipFile(io.BytesIO(xlsx_bytes)) as archive:
            media_files = [name for name in archive.namelist() if name.startswith("xl/media/")]
        self.assertGreaterEqual(len(media_files), 2)

    def test_generate_bon_pdf_adds_confidence_page_when_requested(self):
        from bons.pdf_service import generate_bon_pdf

        extracted = self.new_receipt.extracted_fields
        extracted.final_document_type = "receipt"
        extracted.final_merchant = "RONA"
        extracted.final_total = Decimal("19.49")
        extracted.final_confidence_scores = {
            "merchant_name": 8,
            "total": 9,
        }
        extracted.save(update_fields=[
            "final_document_type",
            "final_merchant",
            "final_total",
            "final_confidence_scores",
        ])

        pdf_bytes = generate_bon_pdf(self.new_bon, include_ai_confidence=True)
        self.assertGreaterEqual(pdf_bytes.count(b"/Type /Page"), 4)

    def test_generate_bon_xlsx_adds_confidence_sheet_when_requested(self):
        import io
        from openpyxl import load_workbook
        from bons.pdf_service import generate_bon_xlsx

        extracted = self.new_receipt.extracted_fields
        extracted.final_document_type = "receipt"
        extracted.final_merchant = "RONA"
        extracted.final_total = Decimal("19.49")
        extracted.final_confidence_scores = {
            "merchant_name": 8,
            "total": 9,
        }
        extracted.save(update_fields=[
            "final_document_type",
            "final_merchant",
            "final_total",
            "final_confidence_scores",
        ])

        xlsx_bytes = generate_bon_xlsx(self.new_bon, include_ai_confidence=True)
        workbook = load_workbook(io.BytesIO(xlsx_bytes))
        self.assertIn("Confiance IA", workbook.sheetnames)
        confidence_sheet = workbook["Confiance IA"]
        values = [
            str(value)
            for row in confidence_sheet.iter_rows(values_only=True)
            for value in row
            if value is not None
        ]
        joined_values = " ".join(values)
        self.assertIn("Confiance IA", joined_values)
        self.assertIn("Marchand", joined_values)
        self.assertIn("RONA", joined_values)


class AuditTrailTests(DuplicateDetectionBaseTest):
    """Test that edits create audit log entries."""

    def test_edit_creates_audit_entry(self):
        from audits.models import AuditLogEntry

        self.client.login(username="tresorier", password="test123")

        # Set bon to editable status
        BonDeCommande.objects.filter(pk=self.existing_bon.pk).update(
            status=BonStatus.READY_FOR_VALIDATION,
        )
        self.existing_bon.refresh_from_db()

        resp = self.client.post(f"/bons/{self.existing_bon.pk}/edit/", {
            "budget_year": self.budget_year.pk,
            "purchase_date": "2026-01-08",
            "short_description": "Updated description",
            "total": "25.00",
            "sub_budget": self.sub_budget.pk,
            "purchaser_member": self.member.pk,
        })
        self.assertEqual(resp.status_code, 302)

        entries = AuditLogEntry.objects.filter(
            target_app_label="bons",
            target_model="bondecommande",
            target_object_id=str(self.existing_bon.pk),
            action="bon.edited",
        )
        self.assertTrue(entries.exists())
        entry = entries.first()
        self.assertIn("short_description", entry.summary)

    def test_no_audit_entry_when_nothing_changed(self):
        from audits.models import AuditLogEntry

        self.client.login(username="tresorier", password="test123")

        BonDeCommande.objects.filter(pk=self.existing_bon.pk).update(
            status=BonStatus.READY_FOR_VALIDATION,
        )
        self.existing_bon.refresh_from_db()

        # Submit the form with the same values
        resp = self.client.post(f"/bons/{self.existing_bon.pk}/edit/", {
            "budget_year": self.budget_year.pk,
            "purchase_date": str(self.existing_bon.purchase_date),
            "short_description": self.existing_bon.short_description,
            "total": str(self.existing_bon.total),
            "sub_budget": self.sub_budget.pk,
            "purchaser_member": self.member.pk,
        })
        self.assertEqual(resp.status_code, 302)

        entries = AuditLogEntry.objects.filter(
            target_app_label="bons",
            target_model="bondecommande",
            target_object_id=str(self.existing_bon.pk),
            action="bon.edited",
        )
        self.assertFalse(entries.exists())


class DuplicateFlagResolveTests(DuplicateDetectionBaseTest):
    """Test duplicate flag resolution."""

    def setUp(self):
        super().setUp()
        self.flag = DuplicateFlag.objects.create(
            receipt_file=self.new_receipt,
            suspected_duplicate_receipt=self.existing_receipt,
            confidence=Decimal("0.95"),
            gpt_comparison_result="Same merchant and amount",
            status=DuplicateFlagStatus.PENDING,
        )

    def test_dismiss_duplicate_flag(self):
        self.client.login(username="tresorier", password="test123")
        resp = self.client.post(
            f"/bons/duplicates/{self.flag.pk}/resolve/",
            {"action": "dismiss"},
        )
        self.assertEqual(resp.status_code, 302)
        self.flag.refresh_from_db()
        self.assertEqual(self.flag.status, DuplicateFlagStatus.DISMISSED)
        self.assertIsNotNone(self.flag.resolved_at)

    def test_confirm_duplicate_flag(self):
        self.client.login(username="tresorier", password="test123")
        resp = self.client.post(
            f"/bons/duplicates/{self.flag.pk}/resolve/",
            {"action": "confirm"},
        )
        self.assertEqual(resp.status_code, 302)
        self.flag.refresh_from_db()
        self.assertEqual(self.flag.status, DuplicateFlagStatus.CONFIRMED_DUPLICATE)


class DetailViewDuplicateTests(DuplicateDetectionBaseTest):
    """Test that the detail view shows duplicate warnings and export gating."""

    def test_detail_shows_duplicate_flag(self):
        DuplicateFlag.objects.create(
            receipt_file=self.new_receipt,
            suspected_duplicate_receipt=self.existing_receipt,
            confidence=Decimal("0.95"),
            gpt_comparison_result="Same merchant",
            status=DuplicateFlagStatus.PENDING,
        )
        self.client.login(username="tresorier", password="test123")
        resp = self.client.get(f"/bons/{self.new_bon.pk}/")
        self.assertContains(resp, "DOUBLON POSSIBLE")

    def test_detail_hides_dismissed_flags(self):
        DuplicateFlag.objects.create(
            receipt_file=self.new_receipt,
            suspected_duplicate_receipt=self.existing_receipt,
            confidence=Decimal("0.95"),
            gpt_comparison_result="Same merchant",
            status=DuplicateFlagStatus.DISMISSED,
        )
        self.client.login(username="tresorier", password="test123")
        resp = self.client.get(f"/bons/{self.new_bon.pk}/")
        self.assertNotContains(resp, "DOUBLON POSSIBLE")

    def test_detail_hides_flags_pointing_to_scan_sessions(self):
        temp_bon = BonDeCommande.objects.create(
            house=self.house,
            budget_year=self.budget_year,
            number="BB260099",
            purchase_date=date(2026, 2, 1),
            short_description="Temp scan",
            total=Decimal("19.49"),
            sub_budget=self.sub_budget,
            purchaser_member=self.member,
            status=BonStatus.READY_FOR_REVIEW,
            is_scan_session=True,
        )
        temp_receipt = ReceiptFile.objects.create(
            bon_de_commande=temp_bon,
            file=SimpleUploadedFile("BC17186.pdf", b"fake temp image", content_type="application/pdf"),
            original_filename="BC17186.pdf",
            content_type="application/pdf",
            ocr_status=OcrStatus.EXTRACTED,
        )
        DuplicateFlag.objects.create(
            receipt_file=self.new_receipt,
            suspected_duplicate_receipt=temp_receipt,
            confidence=Decimal("0.95"),
            gpt_comparison_result="Same temporary scan",
            status=DuplicateFlagStatus.CONFIRMED_DUPLICATE,
        )
        self.client.login(username="tresorier", password="test123")
        resp = self.client.get(f"/bons/{self.new_bon.pk}/")
        self.assertNotContains(resp, "Same temporary scan")
        self.assertNotContains(resp, "BC BB260099")

    def test_detail_shows_confidence_as_real_percentage(self):
        DuplicateFlag.objects.create(
            receipt_file=self.new_receipt,
            suspected_duplicate_receipt=self.existing_receipt,
            confidence=Decimal("0.95"),
            gpt_comparison_result="Same merchant",
            status=DuplicateFlagStatus.PENDING,
        )
        self.client.login(username="tresorier", password="test123")
        resp = self.client.get(f"/bons/{self.new_bon.pk}/")
        self.assertContains(resp, "Confiance : 95%")
        self.assertNotContains(resp, "Confiance : 1%")

    def test_detail_shows_export_link_when_ready(self):
        self.existing_receipt.ocr_status = OcrStatus.CORRECTED
        self.existing_receipt.save(update_fields=["ocr_status"])
        self.client.login(username="tresorier", password="test123")
        resp = self.client.get(f"/bons/{self.existing_bon.pk}/")
        self.assertContains(resp, "Exporter")
        self.assertContains(resp, f'href="/bons/{self.existing_bon.pk}/export/"')

    def test_export_configure_screen_shows_ai_confidence_option(self):
        self.existing_receipt.ocr_status = OcrStatus.CORRECTED
        self.existing_receipt.save(update_fields=["ocr_status"])
        self.client.login(username="tresorier", password="test123")
        resp = self.client.get(f"/bons/{self.existing_bon.pk}/export/")
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, "Configurer l'export")
        self.assertContains(resp, "Inclure les scores de confiance IA")

    def test_export_configure_redirects_with_ai_confidence_flag(self):
        self.existing_receipt.ocr_status = OcrStatus.CORRECTED
        self.existing_receipt.save(update_fields=["ocr_status"])
        self.client.login(username="tresorier", password="test123")
        resp = self.client.post(
            f"/bons/{self.existing_bon.pk}/export/",
            {
                "export_format": "xlsx",
                "include_ai_confidence": "on",
            },
        )
        self.assertEqual(resp.status_code, 302)
        self.assertEqual(
            resp["Location"],
            f"/bons/{self.existing_bon.pk}/xlsx/?include_ai_confidence=1",
        )

    def test_detail_disables_export_when_not_ready(self):
        from bons.models import ReceiptFile, OcrStatus
        from django.core.files.uploadedfile import SimpleUploadedFile
        fake = SimpleUploadedFile("pending.png", b"data", content_type="image/png")
        ReceiptFile.objects.create(
            bon_de_commande=self.existing_bon,
            file=fake, original_filename="pending.png",
            content_type="image/png",
            ocr_status=OcrStatus.PENDING,
        )
        self.client.login(username="tresorier", password="test123")
        resp = self.client.get(f"/bons/{self.existing_bon.pk}/")
        # Should show greyed-out export buttons
        self.assertContains(resp, "cursor:not-allowed")
