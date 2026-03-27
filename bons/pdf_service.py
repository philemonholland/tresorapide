"""PDF and XLSX export for bons de commande.

Generates a PDF matching the paper bon de commande format,
with attached receipt images as subsequent pages.
"""
import io
import os
from decimal import Decimal

from django.conf import settings
from pdf2image import convert_from_path
from pdf2image.exceptions import (
    PDFInfoNotInstalledError,
    PDFPageCountError,
    PDFPopplerTimeoutError,
    PDFSyntaxError,
    PopplerNotInstalledError,
)
from PIL import Image as PILImage, ImageOps, UnidentifiedImageError

from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import inch, cm
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Spacer, Paragraph, Image as RLImage,
    PageBreak,
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

from .ai_confidence import build_receipt_confidence_summary_rows
from .export_formatting import (
    DEFAULT_EXPORT_NUMBER_FORMAT,
    format_money_text,
    normalize_export_number_format,
)


COOP_NAME = "COOPÉRATIVE D'HABITATION\nDES CANTONS DE L'EST"
COOP_ADDRESS = "548, rue Dufferin, Sherbrooke (Québec) J1H 4N1"
COOP_PHONE = "Tél. : (819) 566-6303  Téléc. : (819) 829-1593"
COOP_EMAIL = "Courriel : chce@reseaucoop.com  www.chce.coop"
RECEIPT_PREVIEW_EXCEPTIONS = (
    FileNotFoundError,
    OSError,
    ValueError,
    UnidentifiedImageError,
    PDFInfoNotInstalledError,
    PDFPageCountError,
    PDFPopplerTimeoutError,
    PDFSyntaxError,
    PopplerNotInstalledError,
)


def _fmt_money(val, *, number_format=DEFAULT_EXPORT_NUMBER_FORMAT):
    """Format money text for export output."""
    return format_money_text(val, number_format=number_format)


def _duplicate_flags_for_bon(bon):
    from .models import DuplicateFlag

    receipt_ids = list(bon.active_receipt_files.values_list("pk", flat=True))
    if not receipt_ids:
        return []
    return list(
        DuplicateFlag.objects.actionable().filter(
            receipt_file_id__in=receipt_ids,
        ).select_related(
            "receipt_file",
            "suspected_duplicate_receipt",
            "suspected_duplicate_receipt__bon_de_commande",
        )
    )


def _bon_ai_confidence_summaries(bon):
    summaries = []
    for receipt in bon.active_receipt_files.order_by("created_at", "pk"):
        rows = build_receipt_confidence_summary_rows(receipt)
        if rows:
            summaries.append((receipt, rows))
    return summaries


def _pil_image_to_png_bytes(image_obj) -> bytes:
    image = image_obj.copy()
    if image.mode != "RGB":
        image = image.convert("RGB")
    buf = io.BytesIO()
    image.save(buf, format="PNG", optimize=True)
    return buf.getvalue()


def _receipt_preview_pages(receipt):
    """Return rendered preview pages as (label, png_bytes)."""
    if not receipt.file:
        return []

    file_path = receipt.file.path
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Fichier introuvable : {receipt.original_filename}")

    base_label = receipt.original_filename or os.path.basename(file_path)
    is_pdf = receipt.content_type == "application/pdf" or base_label.lower().endswith(".pdf")

    if is_pdf:
        pdf_pages = convert_from_path(file_path, dpi=150)
        rendered_pages = []
        for index, page in enumerate(pdf_pages, start=1):
            label = base_label if len(pdf_pages) == 1 else f"{base_label} — page {index}"
            rendered_pages.append((label, _pil_image_to_png_bytes(page)))
        return rendered_pages

    with PILImage.open(file_path) as source_image:
        image = ImageOps.exif_transpose(source_image)
        return [(base_label, _pil_image_to_png_bytes(image))]


def _load_receipt_previews(receipt):
    try:
        return _receipt_preview_pages(receipt), None
    except RECEIPT_PREVIEW_EXCEPTIONS as exc:
        return [], str(exc)


def _scaled_reportlab_image(image_bytes, max_width, max_height):
    image_stream = io.BytesIO(image_bytes)
    img = RLImage(image_stream)
    ratio = min(max_width / img.imageWidth, max_height / img.imageHeight)
    ratio = min(ratio, 1)
    img.drawWidth = img.imageWidth * ratio
    img.drawHeight = img.imageHeight * ratio
    return img


def _scaled_xlsx_image(image_bytes, max_width=320, max_height=420):
    from openpyxl.drawing.image import Image as XLImage

    image = XLImage(io.BytesIO(image_bytes))
    ratio = min(max_width / image.width, max_height / image.height)
    ratio = min(ratio, 1)
    image.width = int(image.width * ratio)
    image.height = int(image.height * ratio)
    return image


def generate_bon_pdf(
    bon,
    *,
    include_ai_confidence: bool = False,
    number_format: str = DEFAULT_EXPORT_NUMBER_FORMAT,
) -> bytes:
    """Generate a PDF for a BonDeCommande matching the paper form layout."""
    number_format = normalize_export_number_format(number_format)
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=letter,
        leftMargin=1.5 * cm, rightMargin=1.5 * cm,
        topMargin=1.5 * cm, bottomMargin=1.5 * cm,
    )

    styles = getSampleStyleSheet()
    style_normal = styles["Normal"]
    style_title = ParagraphStyle(
        "BonTitle", parent=styles["Heading2"],
        alignment=TA_CENTER, fontSize=14, spaceAfter=6,
    )
    style_small = ParagraphStyle(
        "Small", parent=style_normal, fontSize=8, leading=10,
    )
    style_center = ParagraphStyle(
        "Center", parent=style_normal, alignment=TA_CENTER, fontSize=9,
    )
    style_right = ParagraphStyle(
        "Right", parent=style_normal, alignment=TA_RIGHT, fontSize=10,
    )
    style_bold = ParagraphStyle(
        "Bold", parent=style_normal, fontSize=10,
        fontName="Helvetica-Bold",
    )

    elements = []

    # ── Header section ───────────────────────────────────────────────────
    header_left = Paragraph(
        f"<b>{COOP_NAME}</b><br/>"
        f"<font size='7'>{COOP_ADDRESS}<br/>"
        f"{COOP_PHONE}<br/>"
        f"{COOP_EMAIL}</font>",
        style_normal,
    )
    header_right = Paragraph(
        f"<font size='8'>Notre numéro de commande</font><br/>"
        f"<b><font size='16'>No {bon.number}</font></b><br/><br/>"
        f"<font size='9'>Date : <b>{bon.purchase_date.strftime('%Y-%m-%d') if bon.purchase_date else '____'}</b></font><br/>"
        f"<font size='9'>Maison : <b>{bon.house.code if bon.house else '____'}</b></font>",
        style_normal,
    )
    header_table = Table(
        [[header_left, header_right]],
        colWidths=[10 * cm, 8 * cm],
    )
    header_table.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ALIGN", (1, 0), (1, 0), "RIGHT"),
    ]))
    elements.append(header_table)
    elements.append(Spacer(1, 0.5 * cm))

    # ── Title ────────────────────────────────────────────────────────────
    elements.append(Paragraph("<b>BON DE COMMANDE</b>", style_title))
    elements.append(Spacer(1, 0.3 * cm))

    # ── Supplier / reimburse section ─────────────────────────────────────
    purchaser_name = bon.purchaser_name_snapshot or (
        bon.purchaser_member.display_name if bon.purchaser_member else ""
    )
    purchaser_apt = bon.purchaser_unit_snapshot or (
        bon.purchaser_apartment.code if bon.purchaser_apartment else ""
    )
    merchant = bon.merchant_name or ""

    info_data = [
        [
            Paragraph("<font size='8'>Fournisseur / personne à rembourser :</font>", style_normal),
            Paragraph("<font size='8'>Emplacement des travaux ou de livraison</font>", style_normal),
        ],
        [
            Paragraph(f"<font size='9'>Nom : <b>{merchant or purchaser_name}</b></font>", style_normal),
            Paragraph(f"<font size='9'>Maison : <b>{bon.house.code if bon.house else ''}</b></font>", style_normal),
        ],
        [
            Paragraph(f"<font size='9'>Adresse : </font>", style_normal),
            Paragraph(f"<font size='9'>Logement(s) : <b>{purchaser_apt}</b></font>", style_normal),
        ],
    ]
    info_table = Table(info_data, colWidths=[10 * cm, 8 * cm])
    info_table.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LINEBELOW", (0, -1), (-1, -1), 0.5, colors.grey),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 0.3 * cm))

    # ── Warning text ─────────────────────────────────────────────────────
    elements.append(Paragraph(
        "<font size='7'><i>Si la personne à rembourser n'est pas un fournisseur, "
        "une signature d'autorisation d'un autre membre est demandée.</i></font>",
        style_center,
    ))
    elements.append(Paragraph(
        "<font size='9'><b>Montant maximum autorisé : 500 $</b></font>",
        style_center,
    ))
    elements.append(Spacer(1, 0.3 * cm))

    # ── Items table ──────────────────────────────────────────────────────
    # Build items from receipts
    items_header = [
        Paragraph("<b><font color='white'>Qté</font></b>", style_center),
        Paragraph("<b><font color='white'>Description</font></b>", style_center),
        Paragraph("<b><font color='white'>Sous-budget</font></b>", style_center),
        Paragraph("<b><font color='white'>Total</font></b>", style_center),
    ]
    items_data = [items_header]

    receipts = bon.active_receipt_files.order_by("created_at", "pk")
    for r in receipts:
        try:
            ef = r.extracted_fields
            desc = ef.final_merchant or ef.merchant_candidate or r.original_filename
            sb_name = ef.sub_budget.name if ef.sub_budget else ""
            total_val = _fmt_money(
                ef.final_total or ef.total_candidate,
                number_format=number_format,
            )
        except Exception:
            desc = r.original_filename
            sb_name = ""
            total_val = ""
        items_data.append([
            Paragraph("1", style_center),
            Paragraph(f"<font size='9'>{desc}</font>", style_normal),
            Paragraph(f"<font size='9'>{sb_name}</font>", style_normal),
            Paragraph(f"<font size='9'>{total_val}</font>", style_right),
        ])

    # Pad to at least 5 rows
    while len(items_data) < 6:
        items_data.append(["", "", "", ""])

    items_table = Table(
        items_data,
        colWidths=[1.5 * cm, 8 * cm, 5 * cm, 3.5 * cm],
    )
    items_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#333333")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("ALIGN", (0, 0), (0, -1), "CENTER"),
        ("ALIGN", (-1, 0), (-1, -1), "RIGHT"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 1), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 1), (-1, -1), 4),
    ]))
    elements.append(items_table)
    elements.append(Spacer(1, 0.3 * cm))

    # ── Totals section ───────────────────────────────────────────────────
    totals_data = []
    if bon.subtotal:
        totals_data.append([
            "Sous-total :",
            _fmt_money(bon.subtotal, number_format=number_format),
        ])
    if bon.tps:
        totals_data.append(["TPS :", _fmt_money(bon.tps, number_format=number_format)])
    if bon.tvq:
        totals_data.append(["TVQ :", _fmt_money(bon.tvq, number_format=number_format)])
    totals_data.append([
        Paragraph("<b>TOTAL</b>", style_right),
        Paragraph(
            f"<b>{_fmt_money(bon.total, number_format=number_format)}</b>",
            style_right,
        ),
    ])

    totals_table = Table(
        totals_data,
        colWidths=[14.5 * cm, 3.5 * cm],
    )
    totals_table.setStyle(TableStyle([
        ("ALIGN", (0, 0), (0, -1), "RIGHT"),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("LINEABOVE", (1, -1), (1, -1), 1, colors.black),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("BOX", (1, -1), (1, -1), 1, colors.black),
    ]))
    elements.append(totals_table)
    elements.append(Spacer(1, 0.5 * cm))

    # ── Signature section ────────────────────────────────────────────────
    sig_data = [
        [
            Paragraph(f"<font size='9'><b>{purchaser_name}</b></font>", style_normal),
            "",
        ],
        [
            Paragraph("<font size='7'>NOM EN LETTRES MOULÉES</font>", style_normal),
            Paragraph("<font size='7'>SIGNATURE</font>", style_normal),
        ],
        ["", ""],
        [
            Paragraph("<font size='7'>NOM EN LETTRES MOULÉES (autorisation)</font>", style_normal),
            Paragraph("<font size='7'>SIGNATURE</font>", style_normal),
        ],
    ]
    sig_table = Table(sig_data, colWidths=[9 * cm, 9 * cm])
    sig_table.setStyle(TableStyle([
        ("LINEBELOW", (0, 0), (0, 0), 0.5, colors.black),
        ("LINEBELOW", (1, 0), (1, 0), 0.5, colors.black),
        ("LINEBELOW", (0, 2), (0, 2), 0.5, colors.black),
        ("LINEBELOW", (1, 2), (1, 2), 0.5, colors.black),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
    ]))
    elements.append(sig_table)
    elements.append(Spacer(1, 0.3 * cm))

    # ── Validation info ──────────────────────────────────────────────────
    if bon.validated_by:
        validator_name = bon.validated_by.get_full_name() or bon.validated_by.username
        validated_text = f"Validé par : {validator_name}"
        if bon.validated_at:
            validated_text += f" — {bon.validated_at.strftime('%Y-%m-%d %H:%M')}"
        elements.append(Paragraph(
            f"<font size='9'><b>{validated_text}</b></font>",
            style_normal,
        ))
        elements.append(Spacer(1, 0.3 * cm))

    # ── Duplicate warning ────────────────────────────────────────────────
    active_dup_flags = _duplicate_flags_for_bon(bon)
    if active_dup_flags:
        style_dup = ParagraphStyle(
            "DupWarning", parent=style_normal,
            fontSize=12, fontName="Helvetica-Bold",
            textColor=colors.red, alignment=TA_CENTER,
            spaceAfter=6,
        )
        elements.append(Spacer(1, 0.3 * cm))
        elements.append(Paragraph(
            "⚠ DOUBLON POSSIBLE ⚠",
            style_dup,
        ))
        for flag in active_dup_flags:
            dup_text = (
                f"Le reçu « {flag.receipt_file.original_filename} » "
                f"correspond possiblement à "
                f"« {flag.suspected_duplicate_receipt.original_filename} » "
                f"(BC {flag.suspected_duplicate_receipt.bon_de_commande.number}) — "
                f"Confiance : {flag.confidence_percent:.0f}%"
            )
            elements.append(Paragraph(
                f"<font size='9' color='red'><b>{dup_text}</b></font>",
                style_normal,
            ))
        elements.append(Spacer(1, 0.3 * cm))

    # ── Footer ───────────────────────────────────────────────────────────
    footer_data = [[
        Paragraph("<font size='7'>Copie blanche : Fournisseur</font>", style_center),
        Paragraph("<font size='7'>Copie jaune : Coopérative</font>", style_center),
        Paragraph("<font size='7'>Copie rose : Maison</font>", style_center),
    ]]
    footer_table = Table(footer_data, colWidths=[6 * cm, 6 * cm, 6 * cm])
    footer_table.setStyle(TableStyle([
        ("BOX", (0, 0), (-1, -1), 0.5, colors.grey),
        ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f0f0f0")),
    ]))
    elements.append(footer_table)

    # ── Attached receipt previews ────────────────────────────────────────
    for receipt in receipts:
        preview_pages, preview_error = _load_receipt_previews(receipt)
        if preview_error:
            elements.append(PageBreak())
            elements.append(Paragraph(
                f"<b>Reçu : {receipt.original_filename}</b>",
                style_bold,
            ))
            elements.append(Paragraph(
                f"<i>(Aperçu non disponible : {preview_error})</i>",
                style_small,
            ))
            continue

        for page_label, preview_bytes in preview_pages:
            elements.append(PageBreak())
            elements.append(Paragraph(
                f"<b>Reçu : {page_label}</b>",
                style_bold,
            ))
            elements.append(Spacer(1, 0.3 * cm))
            elements.append(_scaled_reportlab_image(preview_bytes, 17 * cm, 22 * cm))

    # ── Duplicate comparison pages ───────────────────────────────────────
    if active_dup_flags:
        style_dup_page = ParagraphStyle(
            "DupPageTitle", parent=style_normal,
            fontSize=12, fontName="Helvetica-Bold",
            textColor=colors.red, alignment=TA_CENTER,
            spaceAfter=8,
        )
        style_dup_label = ParagraphStyle(
            "DupLabel", parent=style_normal,
            fontSize=9, fontName="Helvetica-Bold",
            alignment=TA_CENTER,
        )

        for flag in active_dup_flags:
            current_previews, current_error = _load_receipt_previews(flag.receipt_file)
            duplicate_previews, duplicate_error = _load_receipt_previews(
                flag.suspected_duplicate_receipt
            )
            current_number = flag.receipt_file.bon_de_commande.number
            duplicate_number = flag.suspected_duplicate_receipt.bon_de_commande.number

            elements.append(PageBreak())
            elements.append(Paragraph(
                "DOUBLON POSSIBLE — COMPARAISON VISUELLE",
                style_dup_page,
            ))
            elements.append(Paragraph(
                (
                    f"<font size='9'><b>Facture du BC exporté :</b> "
                    f"{flag.receipt_file.original_filename} (BC {current_number})<br/>"
                    f"<b>Facture suspecte :</b> "
                    f"{flag.suspected_duplicate_receipt.original_filename} "
                    f"(BC {duplicate_number})<br/>"
                    f"<b>Confiance :</b> {flag.confidence_percent:.0f}%</font>"
                ),
                style_normal,
            ))
            elements.append(Spacer(1, 0.3 * cm))

            comparison_row = [
                Paragraph("<font color='black'><b>BC exporté</b></font>", style_dup_label),
                Paragraph("<font color='red'><b>POSSIBLE DOUBLON</b></font>", style_dup_label),
            ]
            if current_previews:
                current_block = _scaled_reportlab_image(current_previews[0][1], 8 * cm, 10 * cm)
            else:
                current_block = Paragraph(
                    f"<i>Aperçu non disponible : {current_error or 'non pris en charge'}</i>",
                    style_small,
                )
            if duplicate_previews:
                duplicate_block = _scaled_reportlab_image(
                    duplicate_previews[0][1], 8 * cm, 10 * cm
                )
            else:
                duplicate_block = Paragraph(
                    f"<i>Aperçu non disponible : {duplicate_error or 'non pris en charge'}</i>",
                    style_small,
                )
            compare_table = Table(
                [
                    comparison_row,
                    [current_block, duplicate_block],
                ],
                colWidths=[8.5 * cm, 8.5 * cm],
            )
            compare_table.setStyle(TableStyle([
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("ALIGN", (0, 0), (-1, 0), "CENTER"),
                ("BOX", (0, 0), (-1, -1), 0.5, colors.grey),
                ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ]))
            elements.append(compare_table)

            for page_label, preview_bytes in duplicate_previews[1:]:
                elements.append(PageBreak())
                elements.append(Paragraph(
                    f"POSSIBLE DOUBLON — {page_label}",
                    style_dup_page,
                ))
                elements.append(Spacer(1, 0.2 * cm))
                elements.append(_scaled_reportlab_image(preview_bytes, 17 * cm, 22 * cm))

    if include_ai_confidence:
        confidence_summaries = _bon_ai_confidence_summaries(bon)
        for receipt, rows in confidence_summaries:
            elements.append(PageBreak())
            elements.append(Paragraph(
                f"CONFIANCE IA — {receipt.original_filename}",
                style_title,
            ))
            elements.append(Spacer(1, 0.2 * cm))
            elements.append(Paragraph(
                "<font size='8'>Les scores 0-9 indiquent le niveau de confiance de l'IA. "
                "NA signifie que l'information n'a pas été trouvée sur le document.</font>",
                style_small,
            ))
            elements.append(Spacer(1, 0.3 * cm))

            confidence_rows = [[
                Paragraph("<b>Champ</b>", style_bold),
                Paragraph("<b>Valeur</b>", style_bold),
                Paragraph("<b>Confiance IA</b>", style_bold),
            ]]
            for row in rows:
                confidence_rows.append(
                    [
                        Paragraph(row["label"], style_normal),
                        Paragraph(str(row["value"]), style_normal),
                        Paragraph(
                            f'{row["confidence"]["display"]} - {row["confidence"]["tooltip"]}',
                            style_normal,
                        ),
                    ]
                )

            confidence_table = Table(
                confidence_rows,
                colWidths=[5.5 * cm, 7.5 * cm, 5 * cm],
            )
            confidence_table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#d8edff")),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]))
            elements.append(confidence_table)

    doc.build(elements)
    return buf.getvalue()


def generate_bon_xlsx(
    bon,
    *,
    include_ai_confidence: bool = False,
    number_format: str = DEFAULT_EXPORT_NUMBER_FORMAT,
) -> bytes:
    """Generate an XLSX file for a BonDeCommande."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    number_format = normalize_export_number_format(number_format)
    wb = Workbook()
    ws = wb.active
    ws.title = f"BC {bon.number}"

    # Column widths
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 15

    bold = Font(bold=True)
    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
    border_thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    # Header
    ws.merge_cells("A1:D1")
    ws["A1"] = COOP_NAME.replace("\n", " ")
    ws["A1"].font = bold

    ws.merge_cells("A2:D2")
    ws["A2"] = COOP_ADDRESS

    ws.merge_cells("A4:D4")
    ws["A4"] = "BON DE COMMANDE"
    ws["A4"].font = title_font
    ws["A4"].alignment = Alignment(horizontal="center")

    # Info fields
    row = 6
    info = [
        ("Numéro :", bon.number),
        ("Date :", bon.purchase_date.strftime("%Y-%m-%d") if bon.purchase_date else ""),
        ("Maison :", bon.house.code if bon.house else ""),
        ("Marchand :", bon.merchant_name or ""),
        ("Personne à rembourser :", bon.purchaser_name_snapshot or (
            bon.purchaser_member.display_name if bon.purchaser_member else ""
        )),
        ("Appartement :", bon.purchaser_unit_snapshot or (
            bon.purchaser_apartment.code if bon.purchaser_apartment else ""
        )),
        ("Sous-budget :", bon.sub_budget.name if bon.sub_budget else ""),
    ]
    for label, value in info:
        ws.cell(row=row, column=1, value=label).font = bold
        ws.cell(row=row, column=2, value=value)
        row += 1

    row += 1

    # Items header
    headers = ["Qté", "Description", "Sous-budget", "Total"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border_thin
        cell.alignment = Alignment(horizontal="center")
    row += 1

    # Items from receipts
    receipts = bon.active_receipt_files.order_by("created_at", "pk")
    for r in receipts:
        try:
            ef = r.extracted_fields
            desc = ef.final_merchant or ef.merchant_candidate or r.original_filename
            sb_name = ef.sub_budget.name if ef.sub_budget else ""
            total_amount = ef.final_total or ef.total_candidate or Decimal("0")
        except Exception:
            desc = r.original_filename
            sb_name = ""
            total_amount = Decimal("0")

        ws.cell(row=row, column=1, value=1).border = border_thin
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=2, value=desc).border = border_thin
        ws.cell(row=row, column=3, value=sb_name).border = border_thin
        cell = ws.cell(
            row=row,
            column=4,
            value=format_money_text(total_amount, number_format=number_format),
        )
        cell.border = border_thin
        cell.alignment = Alignment(horizontal="right")
        row += 1

    row += 1

    # Totals
    total_labels = []
    if bon.subtotal:
        total_labels.append(("Sous-total :", bon.subtotal))
    if bon.tps:
        total_labels.append(("TPS :", bon.tps))
    if bon.tvq:
        total_labels.append(("TVQ :", bon.tvq))
    total_labels.append(("TOTAL :", bon.total or Decimal("0")))

    for label, val in total_labels:
        ws.cell(row=row, column=3, value=label).font = bold
        ws.cell(row=row, column=3).alignment = Alignment(horizontal="right")
        cell = ws.cell(
            row=row,
            column=4,
            value=format_money_text(val, number_format=number_format),
        )
        cell.font = bold
        cell.alignment = Alignment(horizontal="right")
        if label == "TOTAL :":
            cell.border = Border(
                top=Side(style="double"), bottom=Side(style="double"),
                left=Side(style="thin"), right=Side(style="thin"),
            )
        row += 1

    row += 2

    # Signatures
    ws.cell(row=row, column=1, value="NOM EN LETTRES MOULÉES :").font = bold
    ws.cell(row=row, column=2, value=bon.purchaser_name_snapshot or (
        bon.purchaser_member.display_name if bon.purchaser_member else ""
    ))
    row += 1
    ws.cell(row=row, column=1, value="SIGNATURE :").font = bold
    row += 2
    ws.cell(row=row, column=1, value="AUTORISATION :").font = bold
    row += 2

    # ── Duplicate warning ────────────────────────────────────────────
    active_dup_flags = _duplicate_flags_for_bon(bon)
    if active_dup_flags:
        dup_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
        dup_font = Font(bold=True, color="CC0000", size=12)
        dup_detail_font = Font(bold=True, color="CC0000", size=9)

        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        cell = ws.cell(row=row, column=1, value="⚠ DOUBLON POSSIBLE ⚠")
        cell.font = dup_font
        cell.fill = dup_fill
        cell.alignment = Alignment(horizontal="center")
        row += 1

        for flag in active_dup_flags:
            dup_text = (
                f"« {flag.receipt_file.original_filename} » ↔ "
                f"« {flag.suspected_duplicate_receipt.original_filename} » "
                f"(BC {flag.suspected_duplicate_receipt.bon_de_commande.number}) — "
                f"Confiance : {flag.confidence_percent:.0f}%"
            )
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
            cell = ws.cell(row=row, column=1, value=dup_text)
            cell.font = dup_detail_font
            cell.fill = dup_fill
            row += 1

        dup_ws = wb.create_sheet("Doublons possibles")
        for col in ["A", "B", "C", "D", "F", "G", "H", "I"]:
            dup_ws.column_dimensions[col].width = 18
        dup_ws.column_dimensions["E"].width = 4

        dup_ws.merge_cells("A1:I1")
        dup_ws["A1"] = "DOUBLONS POSSIBLES — COMPARAISON VISUELLE"
        dup_ws["A1"].font = dup_font
        dup_ws["A1"].fill = dup_fill
        dup_ws["A1"].alignment = Alignment(horizontal="center")

        row = 3
        for index, flag in enumerate(active_dup_flags, start=1):
            current_previews, current_error = _load_receipt_previews(flag.receipt_file)
            duplicate_previews, duplicate_error = _load_receipt_previews(
                flag.suspected_duplicate_receipt
            )
            dup_ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
            dup_ws.cell(row=row, column=1, value=f"POSSIBLE DOUBLON #{index}").font = dup_font
            dup_ws.cell(row=row, column=1).fill = dup_fill
            row += 1

            dup_text = (
                f"Reçu exporté : {flag.receipt_file.original_filename} (BC {flag.receipt_file.bon_de_commande.number}) | "
                f"Reçu suspect : {flag.suspected_duplicate_receipt.original_filename} "
                f"(BC {flag.suspected_duplicate_receipt.bon_de_commande.number}) | "
                f"Confiance : {flag.confidence_percent:.0f}%"
            )
            dup_ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
            dup_ws.cell(row=row, column=1, value=dup_text).font = dup_detail_font
            dup_ws.cell(row=row, column=1).fill = dup_fill
            row += 2

            max_pages = max(len(current_previews), len(duplicate_previews), 1)
            for page_index in range(max_pages):
                current_preview = (
                    current_previews[page_index] if page_index < len(current_previews) else None
                )
                duplicate_preview = (
                    duplicate_previews[page_index] if page_index < len(duplicate_previews) else None
                )

                dup_ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
                dup_ws.merge_cells(start_row=row, start_column=6, end_row=row, end_column=9)
                dup_ws.cell(
                    row=row,
                    column=1,
                    value=(
                        current_preview[0]
                        if current_preview
                        else f"Aperçu non disponible : {current_error or 'non pris en charge'}"
                    ),
                ).font = bold
                dup_ws.cell(
                    row=row,
                    column=6,
                    value=(
                        f"POSSIBLE DOUBLON — {duplicate_preview[0]}"
                        if duplicate_preview
                        else f"Aperçu non disponible : {duplicate_error or 'non pris en charge'}"
                    ),
                ).font = dup_detail_font
                row += 1

                image_anchor_row = row
                if current_preview:
                    dup_ws.add_image(_scaled_xlsx_image(current_preview[1]), f"A{image_anchor_row}")
                if duplicate_preview:
                    dup_ws.add_image(_scaled_xlsx_image(duplicate_preview[1]), f"F{image_anchor_row}")

                for visual_row in range(image_anchor_row, image_anchor_row + 24):
                    dup_ws.row_dimensions[visual_row].height = 24
                row = image_anchor_row + 25

    if include_ai_confidence:
        confidence_summaries = _bon_ai_confidence_summaries(bon)
        if confidence_summaries:
            conf_ws = wb.create_sheet("Confiance IA")
            conf_ws.column_dimensions["A"].width = 28
            conf_ws.column_dimensions["B"].width = 28
            conf_ws.column_dimensions["C"].width = 12
            conf_ws.column_dimensions["D"].width = 40

            conf_ws.merge_cells("A1:D1")
            conf_ws["A1"] = f"BC {bon.number} - Confiance IA"
            conf_ws["A1"].font = title_font
            conf_ws["A1"].alignment = Alignment(horizontal="center")

            conf_row = 3
            for receipt, rows in confidence_summaries:
                conf_ws.merge_cells(start_row=conf_row, start_column=1, end_row=conf_row, end_column=4)
                conf_ws.cell(conf_row, 1, receipt.original_filename).font = bold
                conf_row += 1

                headers = ["Champ", "Valeur", "Score", "Signification"]
                for col, header in enumerate(headers, 1):
                    cell = conf_ws.cell(row=conf_row, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = border_thin
                conf_row += 1

                for summary_row in rows:
                    conf_ws.cell(row=conf_row, column=1, value=summary_row["label"]).border = border_thin
                    conf_ws.cell(row=conf_row, column=2, value=str(summary_row["value"])).border = border_thin
                    conf_ws.cell(
                        row=conf_row,
                        column=3,
                        value=summary_row["confidence"]["display"],
                    ).border = border_thin
                    conf_ws.cell(
                        row=conf_row,
                        column=4,
                        value=summary_row["confidence"]["tooltip"],
                    ).border = border_thin
                    conf_row += 1

                conf_row += 1

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
